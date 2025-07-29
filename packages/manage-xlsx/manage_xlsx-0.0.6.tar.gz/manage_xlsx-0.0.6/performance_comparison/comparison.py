import time
import openpyxl
import xlsx as manage_xlsx
import os

# --- Test Data ---
DATA = [["Header 1", "Header 2", "Header 3", "Header 4", "Header 5"]] + [
    [f"Data {i}-{j}" for j in range(5)] for i in range(1000)
]
FILE_PATH_OPENPYXL = "test_openpyxl.xlsx"
FILE_PATH_MANAGE_XLSX = "test_manage_xlsx.xlsx"


# --- Performance Measurement Functions ---
def measure_openpyxl():
    """Measures the performance of openpyxl."""
    # 1. Create a new workbook and select the active sheet
    start_time_create = time.time()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    end_time_create = time.time()

    # 2. Write data to the sheet
    start_time_write = time.time()
    for row_data in DATA:
        sheet.append(row_data)
    end_time_write = time.time()

    # 3. Save the workbook
    start_time_save = time.time()
    workbook.save(FILE_PATH_OPENPYXL)
    end_time_save = time.time()

    # 4. Load the workbook
    start_time_load = time.time()
    workbook = openpyxl.load_workbook(FILE_PATH_OPENPYXL)
    sheet = workbook.active
    end_time_load = time.time()

    # 5. Read data from the sheet
    start_time_read = time.time()
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    end_time_read = time.time()

    return {
        "create": end_time_create - start_time_create,
        "write": end_time_write - start_time_write,
        "save": end_time_save - start_time_save,
        "load": end_time_load - start_time_load,
        "read": end_time_read - start_time_read,
    }


def measure_manage_xlsx():
    """Measures the performance of manage-xlsx."""
    # 1. Create a new workbook and add a sheet
    start_time_create = time.time()
    workbook = manage_xlsx.Book()
    sheet = workbook.create_sheet("Sheet1", 0)
    end_time_create = time.time()

    # 2. Write data to the sheet
    start_time_write = time.time()
    for row_data in DATA:
        sheet.append(row_data)
    end_time_write = time.time()

    # 3. Save the workbook
    start_time_save = time.time()
    workbook.copy(FILE_PATH_MANAGE_XLSX)
    end_time_save = time.time()

    # 4. Load the workbook
    start_time_load = time.time()
    workbook = manage_xlsx.load_workbook(FILE_PATH_MANAGE_XLSX)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    end_time_load = time.time()

    # 5. Read data from the sheet
    start_time_read = time.time()
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    end_time_read = time.time()

    return {
        "create": end_time_create - start_time_create,
        "write": end_time_write - start_time_write,
        "save": end_time_save - start_time_save,
        "load": end_time_load - start_time_load,
        "read": end_time_read - start_time_read,
    }


if __name__ == "__main__":
    # --- Run Performance Comparison ---
    results_openpyxl = measure_openpyxl()
    results_manage_xlsx = measure_manage_xlsx()

    # --- Clean up generated files ---
    os.remove(FILE_PATH_OPENPYXL)
    os.remove(FILE_PATH_MANAGE_XLSX)

    # --- Print Results ---
    print("## Performance Comparison: openpyxl vs manage-xlsx")
    print("\n| Operation | openpyxl (seconds) | manage-xlsx (seconds) |")
    print("|---|---|---|")
    for op in results_openpyxl:
        print(f"| {op} | {results_openpyxl[op]:.6f} | {results_manage_xlsx[op]:.6f} |")
