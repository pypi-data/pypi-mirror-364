import os
import tempfile
from contextlib import contextmanager

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def modified_excel_data(**data):
    """
    修改excel文件中的单个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "item_name": "A1",
    "item_value": "value"
    }
    """
    try:
        wb = load_workbook(data['file'])
        sheet = wb[data['sheet_name']]
        sheet[data['item_name']] = data['item_value']
        wb.save(data['file'])
    except InvalidFileException:
        raise InvalidFileException("无法打开文件：文件格式不正确或已损坏。")
    except FileNotFoundError:
        raise FileNotFoundError("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")


def modified_excel_datas(**data):
    """
    修改excel文件中的多个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "modified_data": {
        "A1": "value1",
        "B1": "value2",
        "C1": "value3"
    }
    }
    """
    try:
        wb = load_workbook(data['file'])
        sheet = wb[data['sheet_name']]
        modified_data = data['modified_data']
        for key, value in modified_data.items():
            sheet[key] = value
        wb.save(data['file'])
    except InvalidFileException:
        raise InvalidFileException("无法打开文件：文件格式不正确或已损坏。")
    except FileNotFoundError:
        raise FileNotFoundError("文件未找到，请检查文件路径是否正确。")
    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")


@contextmanager
def download_excel(response, filename: str = "downloaded_file.xlsx"):
    """
    上下文管理器：保存 Excel 响应内容到本地，并在退出时自动清理

    :param response: `requests` 的响应对象，需确保 response.content 是 Excel 文件
    :param filename: 自定义保存的文件名（可选，默认 "downloaded_file.xlsx"）
    :return: 返回本地 Excel 文件路径
    """
    if response.status_code != 200:
        raise RuntimeError(f"下载失败，状态码：{response.status_code}")

    # 获取当前项目的临时目录
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)

    # 保存文件
    with open(file_path, "wb") as f:
        f.write(response.content)

    try:
        yield file_path  # 返回 Excel 文件路径
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)  # 退出时自动删除文件

#
# def read_excel_rows(file_path: str, num_rows: int = 1, sheet_name: str = "", read_all: bool = False,
#                     allow_empty: bool = False):
#     """
#     读取 Excel 文件的指定行数据
#     :param file_path: Excel 文件路径
#     :param num_rows: 需要读取的行数（正数表示前几行，负数表示最后几行，默认 1 行）
#     :param sheet_name: 指定的 sheet 名称（为空时读取第一个 sheet）
#     :param read_all: 是否读取所有行数据（默认为 False，读取指定行数）
#     :param allow_empty: 是否允许读取空的 Excel 文件（默认为 False，不允许）
#     :return: 指定的数据列表
#     """
#     if not os.path.exists(file_path):
#         raise FileNotFoundError(f"文件未找到：{file_path}")
#     try:
#         if file_path.endswith(".csv"):
#             # 读取 CSV 文件
#             df = pd.read_csv(file_path, header=None)
#         else:
#             # 如果 sheet_name 为空，则读取第一个 sheet
#             df = pd.read_excel(file_path, sheet_name=sheet_name or 0, header=None)
#         if allow_empty and df.empty:
#             return []
#         if df.empty and not allow_empty:
#             raise ValueError("Excel 文件为空")
#         if read_all:
#             return df.iloc[:].values.tolist()
#         if num_rows < 0:
#             # 返回最后 abs(num_rows) 行数据
#             return df.iloc[num_rows:].values.tolist()
#         else:
#             # 返回前 num_rows 行数据
#             num_rows = min(num_rows, len(df))  # 防止索引超出范围
#             return df.iloc[:num_rows].values.tolist()
#     except Exception as e:
#         raise Exception(f"读取 Excel 文件时发生错误：{e}")

def read_excel_rows(file_path: str, num_rows: int = 1, sheet_name: str = "", read_all: bool = False,
                    allow_empty: bool = False):
    """
    读取 Excel 文件的指定行数据
    :param file_path: Excel 文件路径
    :param num_rows: 需要读取的行数（正数表示前几行，负数表示最后几行，默认 1 行）
    :param sheet_name: 指定的 sheet 名称（为空时读取第一个 sheet）
    :param read_all: 是否读取所有行数据（默认为 False，读取指定行数）
    :param allow_empty: 是否允许读取空的 Excel 文件（默认为 False，不允许）
    :return: 指定的数据列表
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件未找到：{file_path}")
    try:
        # 根据文件类型读取数据
        df = pd.read_csv(file_path, header=None) if file_path.endswith(".csv") else pd.read_excel(
            file_path, sheet_name=sheet_name or 0, header=None)

        if df.empty:
            if allow_empty:
                return []
            raise ValueError("Excel 文件为空")

        if read_all:
            return df.values.tolist()

        return df.tail(abs(num_rows)).values.tolist() if num_rows < 0 else df.head(num_rows).values.tolist()
    except Exception as e:
        raise Exception(f"读取 Excel 文件时发生错误：{e}")

def get_sheet_names(file_path):
    """ 读取 Excel 文件中的所有 sheet 名称 """
    with pd.ExcelFile(file_path) as xls:
        return xls.sheet_names


def save_excel_datas(**data):
    """
    重新保存excel文件中的多个数据
    data = {
    "file": "文件路径",
    "sheet_name": "表名",
    "modified_data": {
        "A1": "value1",
        "B1": "value2",
        "C1": "value3"
    }
    }
    """
    import xlsxwriter
    try:
        # 创建新的工作簿
        workbook = xlsxwriter.Workbook(data['file'])
        sheet = workbook.add_worksheet(data['sheet_name'])

        # 修改数据
        modified_data = data['modified_data']
        for cell, value in modified_data.items():
            col = ord(cell[0].upper()) - ord('A')
            row = int(cell[1:]) - 1
            sheet.write(row, col, value)

        # 保存文件
        workbook.close()

    except Exception as e:
        raise Exception(f"修改excel文件遇到未知错误：{e}")


def modified_excel_datas_xls(**data):
    """
    修改xls文件中的多个数据
    data = {
        "file": "文件路径",
        "sheet_name": "表名",
        "modified_data": {
            "0_0": "value1",  # 格式为 "行_列"，例如 "0_0" 表示第 0 行第 0 列
            "0_1": "value2",
            "1_0": "value3"
        }
    }
    """
    from xlrd import open_workbook
    from xlutils.copy import copy
    try:
        # 打开原始工作簿
        rb = open_workbook(data['file'], formatting_info=True)

        # 检查是否存在指定的工作表
        if data['sheet_name'] not in rb.sheet_names():
            raise ValueError(f"Sheet '{data['sheet_name']}' 不存在于文件中。")

        # 获取工作表索引
        sheet_index = rb.sheet_names().index(data['sheet_name'])

        # 复制工作簿以便修改
        wb = copy(rb)
        sheet = wb.get_sheet(sheet_index)

        # 获取需要修改的数据
        modified_data = data['modified_data']

        # 遍历字典并更新单元格
        for position, value in modified_data.items():
            row, col = map(int, position.split('_'))
            sheet.write(row, col, value)

        # 保存工作簿
        # 创建临时文件名
        temp_file = data['file'] + '.tmp'
        wb.save(temp_file)

        # 替换原文件
        os.remove(data['file'])  # 删除原文件
        os.rename(temp_file, data['file'])  # 将新文件重命名为原文件名

    except Exception as e:
        raise Exception(f"修改xls文件遇到错误：{e}")


def clear_excel_rows(**data):
    """
    清空 Excel 文件中所有工作表的第二行及之后的所有行
    data={
        "file": "文件路径",
        “sheet_name": "表名",  # 可选，如果不指定则清空所有工作表
        "start_row":   # 删除从第几行开始删除
    }
    """
    clear_sheet = data.get('sheet_name')
    workbook = load_workbook(data['file'])
    for sheet_name in workbook.sheetnames:
        if clear_sheet and sheet_name != clear_sheet:
            continue
        worksheet = workbook[sheet_name]
        if worksheet.max_row > data['start_row']:  # 确保有数据行才删除
            worksheet.delete_rows(data['start_row'], worksheet.max_row - 1)
    workbook.save(data['file'])  # 保存清空后的文件
