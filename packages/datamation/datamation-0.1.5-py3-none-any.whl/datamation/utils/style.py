#!/usr/bin/env python
# -*- coding: utf-8 -*-
import time
__all__ = ["excel","HTML_ZEBRA"]

# 灰色样式的css
HTML_ZEBRA = '''<style>
                table.dataframe{
                    border: 1px solid #888888;
                    border-collapse: collapse;
                    font-family: Arial,Helvetica,sans-serif;
                    margin-top: 10px;
                    width: 100%;
                }
                table.dataframe th {
                    background-color: #CCCCCC;
                    border: 1px solid #888888;
                    padding: 5px 15px 5px 5px;
                    text-align: left;
                    vertical-align: baseline;
                }
                table.dataframe td {
                    background-color: #EFEFEF;
                    border: 1px solid #AAAAAA;
                    padding: 5px 15px 5px 5px;
                    vertical-align: text-top;
                }
                </style>'''



def excel(fpath = None):
    '''配置excel的报表样式.

    Parameters
    -----------

    fpath : str
        excel文件的绝对路径

    '''
    from openpyxl import load_workbook
    from openpyxl.styles import Border,PatternFill,alignment,Side,Alignment,Font,Fill
    import pandas as pd

    if fpath:
        '''配置excel报表的样式
        fpath : excel文件对象的路径和名称'''
        print(time.ctime(),'开始样式配置...')
        wb = load_workbook(filename = fpath)
        # 通用样式
        line_t = Side(style='thin', color='000000')  
        border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)
        # 数据表样式设置
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            # 细边框线
            def sty1(x):
                x.fill = PatternFill('solid', fgColor='D9D9D9')
                x.border =border
                x.alignment=Alignment(horizontal='center', vertical='center')
                return x
            def sty2(x):
                x.fill = PatternFill('solid', fgColor='FFFFFF')
                x.border =border
                x.alignment=Alignment(horizontal='center', vertical='center')
                return x
            df = pd.DataFrame([n for n in ws.rows])
            df[::2].applymap(sty1)
            df[1::2].applymap(sty2)
        wb.save(fpath)
        #print(time.ctime(),'数据样式配置结束')

    # 报表标题设置
    #print(time.ctime(),'报表主题配置开始...')
    wb = load_workbook(filename = fpath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        ws.cell(1, 1).value = f'{sheet}'
        ws.cell(1, 1).font  = Font(name='微软雅黑', color='FF0000', size=15, b=True)
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(1, 1).fill =PatternFill('solid', fgColor='FFFFFF') # 填充白色
        ws.freeze_panes = ws.cell(3,1)  # 在A3单元格处冻结主题行和数据列标题行
    wb.save(fpath)
    print(time.ctime(),'报表样式配置结束...')

