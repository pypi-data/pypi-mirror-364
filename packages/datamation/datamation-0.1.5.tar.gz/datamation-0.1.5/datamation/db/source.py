import gzip
import csv
import threading
import queue
from itertools import chain

__all__ = ['source_sql','source_pandas','source_csv','source_xlsx']

class Source:
    def __init__(self):
        self.tag_out = None
        self.execute = None
        self.startup = None

    def __and__(self, other):
        return self.unionall(other)

    def _startup(self):
        for row in self:
            self.execute(self.tag_out, row)
        self.tag_out.endload()

    def unionall(self, other):
        if isinstance(other, self.__class__):
            self.__iter__ = chain(self, other)
            return self

class ThreadedSource:
    def __init__(self, arraysize=20000):
        self.arraysize = arraysize*10
        self.data_queue = queue.Queue(maxsize=arraysize*10)
        self._stop_fetch = threading.Event()
        self._fetch_thread = threading.Thread(target=self._data_fetcher, daemon=True)
        self._fetch_thread.start()
        self._iter_done = False  # 标记迭代是否完成

    def _data_fetcher(self):
        """子类应实现的数据获取逻辑"""
        raise NotImplementedError("Subclasses must implement _data_fetcher")

    def __iter__(self):
        """返回自身以支持迭代协议"""
        return self

    def __next__(self):
        """获取下一个数据。如果迭代完成则释放资源。"""
        if not self._iter_done:
            if not self._stop_fetch.is_set() or not self.data_queue.empty():
                try:
                    return self.data_queue.get(timeout=10)
                except queue.Empty:
                    pass
            self._iter_done = True
            self._close_resources()
        raise StopIteration

    def stop_thread(self):
        """通知线程停止"""
        self._stop_fetch.set()
        self._fetch_thread.join()

    def _close_resources(self):
        """子类实现的资源释放逻辑"""
        raise NotImplementedError("Subclasses must implement _close_resources")

    def __del__(self):
        """确保销毁对象时释放资源"""
        if not self._iter_done:
            self._close_resources()
            self.stop_thread()


class source_sql(ThreadedSource):
    
    def __init__(self, conn, sql, params=None, arraysize=20000, rename=None, data_format='dict'):
        self.conn_factory = conn if callable(conn) else None
        self.conn = conn() if self.conn_factory else conn
        self.cursor = self.conn.cursor()
        self.cursor.arraysize = arraysize
        self.data_format = data_format
        self.rename = rename or {}
        self.params = params
        self.sql = sql
        self.rowcount = self._get_rowcount()  # 获取查询总行数
        self._setup_cursor()
        super().__init__(arraysize)

    def _setup_cursor(self):
        """设置游标，执行 SQL 查询。"""
        if self.params:
            self.cursor.execute(self.sql, self.params)
        else:
            self.cursor.execute(self.sql)
        self.cols = [self.rename.get(col[0], col[0]) for col in self.cursor.description]

    def _get_rowcount(self):
        """计算查询总行数"""
        count_query = f"SELECT COUNT(*) FROM ({self.sql}) AS subquery"
        temp_conn = self.conn_factory() if self.conn_factory else self.conn
        temp_cursor = temp_conn.cursor()
        try:
            if self.params:
                temp_cursor.execute(count_query, self.params)
            else:
                temp_cursor.execute(count_query)
            count = temp_cursor.fetchone()[0]
        finally:
            temp_cursor.close()
            if self.conn_factory:
                temp_conn.close()
        return count

    def _data_fetcher(self):
        """从数据库中获取数据并存入队列。"""
        try:
            while not self._stop_fetch.is_set():
                data = self.cursor.fetchmany(self.cursor.arraysize)
                if not data:
                    break
                for row in data:
                    if self.data_format == 'dict':
                        row = dict(zip(self.cols, row))
                    self.data_queue.put(row)
        except Exception as e:
            print(f"Error fetching data: {e}")

    def _close_resources(self):
        """关闭游标和连接"""
        self.cursor.close()
        if self.conn_factory:
            try:
                self.conn.close()
            except Exception as e:
                pass

    def __del__(self):
        """在对象销毁时关闭资源。"""
        self._close_resources()

    def __len__(self):
        return self.rowcount

class source_pandas(Source):
    def __init__(self, dataframe, rename=None):
        super().__init__()
        self.dataframe = dataframe.rename(columns=rename or {}).where(dataframe.notnull(), None)
        self._iter = iter(self.dataframe.to_dict('records'))
        self.cols = list(self.dataframe.columns)

    def __iter__(self):
        return self._iter

class source_xlsx(Source):
    def __init__(self, file, sheet_name=None, rename=None, to_datetime=None):
        from openpyxl import load_workbook
        super().__init__()
        self.ws = load_workbook(filename=file, data_only=True)[sheet_name or load_workbook(file).sheetnames[0]]
        self.rename = rename or {}
        self.to_datetime = to_datetime or []
        self.cols = []
        
        for col in next(self.ws.iter_rows(min_row=1, max_row=1)):
            if col.value:
                self.cols.append(self.rename.get(col.value, col.value))
            else:
                break
        self.max_col = len(self.cols)
        self._iter = self._create_iter()

        if self.to_datetime:
            try:
                global parse
                from dateutil.parser import parse
            except ImportError:
                raise ImportError("Please install dateutil to use this class. You can install it with 'pip install python-dateutil'.")


    def _create_iter(self):
        for row in self.ws.iter_rows(min_row=2,max_col = self.max_col, values_only=True):
            record = dict(zip(self.cols, row))
            for col in self.to_datetime:
                if col in record:
                    record[col] = parse(record[col])
            if any(record.values()):
                yield record
            else:
                continue

    def __iter__(self):
        return self._iter


class source_csv(ThreadedSource):
    def __init__(self, file, fieldnames=None, arraysize=30000, mapping=None, to_datetime=None, delimiter=',', encoding='utf-8'):
        super().__init__(arraysize)
        self.file = gzip.open(file, 'rt', encoding=encoding) if file.endswith('.gz') else open(file, 'rt', encoding=encoding)
        self.reader = csv.DictReader(self.file, fieldnames=fieldnames, delimiter=delimiter)
        self.mapping = mapping or {}
        self.to_datetime = to_datetime or []
        self.cols = fieldnames if fieldnames else self.reader.fieldnames
        if self.to_datetime:
            try:
                global parse
                from dateutil.parser import parse
            except ImportError:
                raise ImportError("Please install dateutil to use this class. You can install it with 'pip install python-dateutil'.")

    def _data_fetcher(self):
        for row in self.reader:
            for col, func in self.mapping.items():
                if col in row:
                    row[col] = func(row[col])
            for col in self.to_datetime:
                if col in row:
                    row[col] = parse(row[col])
            self.data_queue.put(row)

class source_es(ThreadedSource):
    def __init__(self, conn, index, **kwargs):
        try:
            from elasticsearch_dsl import Search,Q
        except ImportError:
            raise ImportError("Please install elasticsearch_dsl to use this class.")
        super().__init__()
        self.search = Search(using=conn, index=index).query('bool', must=[Q(k.lower(), **v) for k, v in kwargs.items()])

    def _data_fetcher(self):
        for row in self.search.scan():
            self.data_queue.put(row.to_dict())