import click
import re, os, io
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from typing import Any, Dict, Union, Literal, Annotated, Optional, Tuple, List
from pydantic import Field
from fastmcp import FastMCP
from zlxmcp.utils.oss import OSS
from zlxmcp.utils.utils import project_cache_dir

mcp = FastMCP("ZLX-MCP")
FILE_METHOD = os.getenv("FILE_METHOD", "local")
print(f"[ZLXMCP] USE {FILE_METHOD} FILE")
if FILE_METHOD == "oss":
    oss_key = os.getenv("OSS_KEY")
    assert oss_key, "请设置环境变量 OSS_KEY"
    oss = OSS(
        access_key_id=oss_key.split("-")[0],
        access_key_secret=oss_key.split("-")[1],
        bucket_name="-".join(oss_key.split("-")[2:]),
    )


@mcp.tool()
def get_tool_version() -> str:
    """获取当前工具的版本号"""
    from zlxmcp import __version__
    return __version__


@mcp.tool()
def list_excel_files() -> Tuple[str, List[str]]:
    """
    获取指定目录下的所有excel文件

    Args:
        path: 路径

    Returns:

    """
    if FILE_METHOD == "oss":
        msg, files = oss.list_objects("zlxmcp/data/")
        files = [f for f in files if f.endswith(".xlsx")]
    elif FILE_METHOD == "local":
        path = os.path.join(project_cache_dir(), "data")
        if not os.path.exists(path):
            os.makedirs(path)
        files = [f for f in os.listdir(path) if f.endswith(".xlsx")]
        msg = "成功获取文件列表"
    else:
        raise ValueError(f"不支持的文件方法: {FILE_METHOD}")
    return msg, files


def get_file_content(file_name: str) -> Tuple[bool, str, Optional[bytes]]:
    """"""
    content = None
    if FILE_METHOD == "oss":
        file_path = f"zlxmcp/data/{file_name}"
        msg, is_exist = oss.object_exist(object_name=file_path)
        if not is_exist:
            msg = f"`{file_path}` 文件不存在"
        else:
            msg, content = oss.download_file(object_name=file_path)
    else:
        path = os.path.join(project_cache_dir(), "data")
        file_path = os.path.join(path, file_name)
        is_exist = os.path.exists(file_path)
        if not is_exist:
            msg = f"`{file_name}` 文件不存在"
        else:
            with open(file_path, "rb") as f:
                content = f.read()
            msg = "成功获取文件内容"
    return is_exist, msg, content


@mcp.tool()
def excel_info(
        file_name: Annotated[Optional[str], Field(description="Excel文件名")] = None,
) -> str:
    """
    获取指定excel文件中的所有sheet名称，以及sheet中的所有列名

    Args:
        file_name: 文件名

    Returns:

    """
    is_exist, msg, content = get_file_content(file_name=file_name)

    if not is_exist:
        return msg

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        result = [f"文件: {file_name}", f"包含的sheet: {', '.join(sheet_names)}"]

        for sheet in sheet_names:
            try:
                # 只读取第一行获取列名
                df = pd.read_excel(io.BytesIO(content), sheet_name=sheet, nrows=1)
                columns = df.columns.tolist()
                result.append(f"\nSheet [{sheet}] 的列名:")
                result.append(", ".join([str(col) for col in columns]))
            except Exception as e:
                result.append(f"\n读取Sheet [{sheet}] 列名时出错: {str(e)}")
        return "\n".join(result)
    except Exception as e:
        return f"读取文件 {file_name} 时出错: {str(e)}"


def _sanitize_code(code: str) -> str:
    """
    对pandas代码进行安全检查和转换

    参数:
        code: 原始代码

    返回:
        安全处理后的代码

    异常:
        ValueError: 如果检测到不安全代码
    """
    # 1. 移除注释和多行字符串
    code = re.sub(r'#.*', '', code)  # 移除单行注释
    code = re.sub(r'""".*?"""', '', code, flags=re.DOTALL)  # 移除多行注释

    # 2. 黑名单检查
    blacklist = [
        'os.', 'subprocess.', 'sys.', 'eval(', 'exec(',
        'open(', 'import ', '__import__', 'lambda',
        'globals(', 'locals(', 'compile(', 'execfile(',
        'getattr', 'setattr', 'delattr', 'property',
        'input(', 'help(', 'memoryview', 'bytearray',
        'reload', '__build_class__', 'super', 'vars('
    ]

    for forbidden in blacklist:
        if forbidden.lower() in code.lower():
            raise ValueError(f"检测到潜在危险代码: {forbidden}")
    return code


def _get_safe_builtins() -> Dict[str, Any]:
    """返回安全的builtins字典"""
    return {
        'None': None,
        'True': True,
        'False': False,
        'bool': bool,
        'int': int,
        'float': float,
        'str': str,
        'list': list,
        'tuple': tuple,
        'dict': dict,
        'set': set,
        'len': len,
        'range': range,
        'sum': sum,
        'min': min,
        'max': max,
        'abs': abs,
        'round': round,
        'zip': zip
    }


@mcp.tool()
def execute_pandas(
        file_name: str,
        sheet_name: Optional[str] = None,
        code: Optional[str] = None,
) -> Tuple[bool, str, Optional[pd.DataFrame]]:
    """执行pandas代码，并返回结果，
        1. `df`为已经读取的Excel数据
        2. 执行结果存储于`df_result`变量中
        3. `code`代码仅需要写对于`df`的操作即可

    Examples:
        # 取出前3行
        df_result = df.head(3)

    Args:
        file_name: Excel文件名
        sheet_name: 工作表名(可选)
        code: 要执行的pandas代码

    Returns:
        Tuple[bool, str, Optional[pd.DataFrame]]:
            (执行状态, 消息, 结果DataFrame)
    """
    is_exist, msg, content = get_file_content(file_name)

    if not is_exist:
        return False, msg, None

    try:
        # 2. 读取Excel文件
        df = pd.read_excel(io.BytesIO(content), sheet_name=sheet_name)

        # 3. 如果没有提供代码，直接返回DataFrame
        if not code:
            return True, f"成功读取文件 {file_name}", df

        # 4. 安全检查和执行代码
        try:
            # 安全检查
            safe_code = _sanitize_code(code)

            # 准备安全执行环境
            global_vars = {
                '__builtins__': _get_safe_builtins(),
                'pd': pd, 'np': np
            }
            local_vars = {"df": df}
            if isinstance(df, Dict):
                return False, "需要指定sheet_name", None

            exec(safe_code, global_vars, local_vars)

            # 获取结果
            df_result = local_vars.get('df_result', None)

            if df_result is not None:
                return True, "代码执行成功", df_result
            else:
                return False, "df_result is Null", None

        except ValueError as ve:
            return False, f"代码安全检查失败: {str(ve)}", None
        except Exception as e:
            return False, f"代码执行出错: {str(e)}", None

    except Exception as e:
        return False, f"读取文件 {file_name} 时出错: {str(e)}", None


@click.command()
@click.option("--host", "-h", default="127.0.0.1", type=str, required=False, help="host")
@click.option("--port", "-p", default=8000, type=int, required=False, help="port")
@click.option("--transport", "-t", default="stdio", type=str, required=False, help="transport")
def pandas_mcp_server(
        host: str = None,
        port: Union[int, str] = None,
        transport: Literal["stdio", "sse", "streamable-http"] = "stdio",
) -> None:
    """"""
    if transport == "sse":
        mcp.run(transport=transport, port=port, host=host)
    else:
        mcp.run(transport=transport)


if __name__ == "__main__":
    mcp.run(transport="sse", port=9003)
