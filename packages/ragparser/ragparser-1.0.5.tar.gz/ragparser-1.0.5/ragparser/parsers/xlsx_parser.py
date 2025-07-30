"""
XLSX parser implementation using openpyxl
"""

import asyncio
import logging
from typing import List, Dict, Any, Optional, TYPE_CHECKING
from pathlib import Path
import io

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    # Create dummy class for type hints
    if TYPE_CHECKING:
        from openpyxl.workbook import Workbook
    else:
        Workbook = object

from .base import BaseParser
from ..core.models import ParsedDocument, ParserConfig, FileType, ContentBlock
from ..core.exceptions import ProcessingError


class XLSXParser(BaseParser):
    """Parser for XLSX documents"""

    def __init__(self):
        super().__init__()
        self.supported_formats = [FileType.XLSX]

        if not HAS_OPENPYXL:
            raise ImportError(
                "XLSX parsing requires openpyxl. Install with: pip install openpyxl"
            )

    async def parse_async(
        self, file_path: Path, config: ParserConfig
    ) -> ParsedDocument:
        """Parse XLSX file asynchronously"""
        try:
            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(
                None, self._parse_xlsx_sync, str(file_path), config
            )
        except Exception as e:
            raise ProcessingError(f"Failed to parse XLSX: {str(e)}", e)

    async def parse_from_bytes_async(
        self, data: bytes, filename: str, config: ParserConfig
    ) -> ParsedDocument:
        """Parse XLSX from bytes asynchronously"""
        try:
            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(
                None, self._parse_xlsx_from_bytes_sync, data, filename, config
            )
        except Exception as e:
            raise ProcessingError(f"Failed to parse XLSX from bytes: {str(e)}", e)

    def _parse_xlsx_sync(self, file_path: str, config: ParserConfig) -> ParsedDocument:
        """Synchronous XLSX parsing"""
        path = Path(file_path)
        document = asyncio.run(
            self._create_base_document(
                path, path.name, FileType.XLSX, path.stat().st_size
            )
        )

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            return self._extract_content(wb, document, config)
        except Exception as e:
            raise ProcessingError(f"Failed to parse XLSX: {str(e)}", e)

    def _parse_xlsx_from_bytes_sync(
        self, data: bytes, filename: str, config: ParserConfig
    ) -> ParsedDocument:
        """Synchronous XLSX parsing from bytes"""
        document = asyncio.run(
            self._create_base_document(None, filename, FileType.XLSX, len(data))
        )

        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            return self._extract_content(wb, document, config)
        except Exception as e:
            raise ProcessingError(f"Failed to parse XLSX from bytes: {str(e)}", e)

    def _extract_content(
        self, wb, document: ParsedDocument, config: ParserConfig
    ) -> ParsedDocument:
        """Extract content from XLSX workbook"""
        """Extract content from XLSX workbook"""
        content_parts = []
        content_blocks = []
        tables = []

        # Extract workbook properties
        if wb.properties:
            props = wb.properties
            document.metadata.title = props.title
            document.metadata.author = props.creator
            document.metadata.subject = props.subject
            document.metadata.creation_date = props.created
            document.metadata.modification_date = props.modified

        # Process each worksheet
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            try:
                sheet = wb[sheet_name]
                sheet_content = []

                # Get sheet dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column

                if max_row == 1 and max_col == 1:
                    # Empty sheet
                    continue

                # Add sheet header
                sheet_header = f"=== Sheet: {sheet_name} ==="
                content_blocks.append(
                    ContentBlock(
                        content=sheet_header,
                        block_type="header",
                        page_number=sheet_idx + 1,
                    )
                )
                sheet_content.append(sheet_header)

                # Extract data as table
                sheet_data = []
                headers = []

                # Process rows
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    if row_idx > 1000:  # Limit to prevent memory issues
                        document.extraction_notes.append(
                            f"Sheet '{sheet_name}' truncated at 1000 rows"
                        )
                        break

                    # Clean row data
                    row_data = []
                    for cell_value in row:
                        if cell_value is None:
                            row_data.append("")
                        else:
                            row_data.append(str(cell_value).strip())

                    # Skip completely empty rows
                    if not any(cell for cell in row_data):
                        continue

                    if row_idx == 1:
                        # First row as headers
                        headers = row_data

                    sheet_data.append(row_data)

                if sheet_data:
                    # Add as table
                    if config.extract_tables:
                        tables.append(
                            {
                                "sheet": sheet_name,
                                "sheet_index": sheet_idx,
                                "headers": headers,
                                "data": sheet_data,
                                "dimensions": {
                                    "rows": len(sheet_data),
                                    "cols": max_col,
                                },
                            }
                        )

                    # Convert to text format
                    table_text = self._format_sheet_as_text(
                        sheet_name, headers, sheet_data, config
                    )

                    content_blocks.append(
                        ContentBlock(
                            content=table_text,
                            block_type="table",
                            page_number=sheet_idx + 1,
                        )
                    )
                    sheet_content.append(table_text)

                # Process named ranges if present
                try:
                    if hasattr(wb, "defined_names"):
                        for named_range in wb.defined_names:
                            if sheet_name in str(named_range.destinations):
                                sheet_content.append(f"Named Range: {named_range.name}")
                except Exception as e:
                    logging.warning(f"Failed to process named ranges: {str(e)}")

                # Add sheet content
                if sheet_content:
                    content_parts.extend(sheet_content)

            except Exception as e:
                logging.warning(f"Failed to process sheet '{sheet_name}': {str(e)}")
                document.extraction_notes.append(
                    f"Failed to process sheet '{sheet_name}': {str(e)}"
                )
                continue

        # Combine content
        document.content = "\n\n".join(content_parts)
        document.content_blocks = content_blocks
        document.tables = tables

        # Update metadata
        document.metadata.page_count = len(wb.sheetnames)
        document.metadata.word_count = len(document.content.split())
        document.metadata.character_count = len(document.content)

        return document

    def _format_sheet_as_text(
        self,
        sheet_name: str,
        headers: List[str],
        data: List[List[str]],
        config: ParserConfig,
    ) -> str:
        """Format sheet data as readable text"""
        lines = []

        if headers and any(headers):
            # Add headers
            lines.append("Headers: " + " | ".join(headers))
            lines.append("-" * 50)

        # Add data rows (limit for readability)
        max_rows_to_show = config.custom_settings.get("max_preview_rows", 20)

        for i, row in enumerate(data[:max_rows_to_show]):
            if i == 0 and headers:
                continue  # Skip header row if already processed

            row_text = " | ".join(str(cell) for cell in row)
            if row_text.strip():
                lines.append(f"Row {i+1}: {row_text}")

        if len(data) > max_rows_to_show:
            lines.append(f"... and {len(data) - max_rows_to_show} more rows")

        return "\n".join(lines)
