# src/file_downloader/server.py
import os, urllib.parse
import httpx
from fastmcp import FastMCP
import openpyxl
import shutil
import json
from openpyxl import load_workbook
from openpyxl.cell import Cell
from copy import copy
from openpyxl.styles import Side, Border
from datetime import datetime
import uuid as uuid_lib

def main():
    mcp = FastMCP("unieai-mcp-excel-stdio")

    """
    @mcp.tool()
    def write_data_to_excel_with_custom(
        filepath: str,
        outputpath: str,
        data: dict,
    ) -> str:
        """"""
        將資料寫入 Excel 模板中對應的 {{key}} 位置。

        範例:
        write_data_to_excel_with_custom(
            filepath="D:/excel_temp.xlsx",
            outputpath="D:/excel_filled.xlsx",
            data={"item": "產品A", "number": "2", "price": 10000}
        )
        """"""
        try:
            params = {
                "filepath": filepath,
                "outputpath": outputpath,
                "data": data
            }
            return fill_excel(params)
        except Exception as e:
            return f"Error: {e}"
    """

    @mcp.tool()
    def write_data_to_excel_with_insert_row(
        data: list[list],
    ) -> str:
        """
        UnieAI專用報價單模板A
        將資料寫入 Excel 模板中，
        將傳入的資料寫入 Excel，第一筆固定位置，
        第二筆開始每寫一筆先插入新列，確保後筆永遠緊接上一筆。
        請幫我確認你傳入的資料是否正確??
        資料不正確的話請再呼叫一次。

        範例:
        write_data_to_excel_with_insert_row(
            data=[
                    [ "-", "產品A", "2", 10000, ""],
                    [ "", "產品B", "33", 555, ""],
                    [ "", "產品C", "44", 666, ""]
                ]
        )
        """
        try:
            outputpath = "/app/data/storage/unieai-mcp-excel/excel_acer_a_" + datetime.now().strftime("%Y%m%d") + "_" + str(uuid_lib.uuid4()) + ".xlsx"
            params = {
                "filepath": "/app/data/storage/unieai-mcp-excel/excel_acer_temp_a.xlsx",
                "outputpath": outputpath,
                "sheet_name": "英文翻譯",
                "data_rows": data
            }
            return fill_excel_with_insert_row(params)
        except Exception as e:
            return f"Error: {e}"

    mcp.run(transport="stdio")


def fill_excel(params: dict) -> str:
    src = params["filepath"]
    dst = params["outputpath"]
    data = params["data"]

    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    sheet_name = "英文翻譯"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到名稱為 \"{sheet_name}\" 的工作表")
    ws = wb[sheet_name]

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{" in cell.value:
                for key, val in data.items():
                    cell.value = cell.value.replace(f"{{{{{key}}}}}", str(val))

    wb.save(dst)
    wb.close()
    return dst




def fill_excel_with_insert_row(params: dict) -> str:
    src = params["filepath"]
    dst = params["outputpath"]
    sheet_name = params["sheet_name"]
    data = params["data_rows"]

    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copy(src, dst)

    wb = openpyxl.load_workbook(dst)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"找不到名稱為 \"{sheet_name}\" 的工作表")
    ws = wb[sheet_name]

    # === 建立「黑色薄邊框」樣式（四邊都細線） ===
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    start_col = 2   # 代表 B 欄
    start_row = 19

    # 假設 data 已先轉成巢狀 list，例如：
    # data = [
    #     ["產品A", 2,  10000],
    #     ["產品B", 22, 2200],
    # ]

    for i, row_vals in enumerate(data):
        target_row = start_row + i
        if i > 0:                         # 第二筆起，先插入一列
            ws.insert_rows(target_row)

        for j, val in enumerate(row_vals):
            cell = ws.cell(row=target_row, column=start_col + j)
            cell.value = val
            cell.border = border_all      # ★ 套用邊框（關鍵一行）




    # start_col = 3   # C 欄
    # start_row = 20

    # for i, row_vals in enumerate(data):
    #     target_row = start_row + i
    #     if i > 0:                      # 第二筆起，先插入一列
    #         ws.insert_rows(target_row)

    #     for j, val in enumerate(row_vals):
    #         ws.cell(row=target_row, column=start_col + j, value=val)


    link_url = "https://office-mcp-dl.unieai.com/unieai-mcp-excel/"+dst.split("/")[-1]
    #response = "請幫我確認你傳入的資料是否正確?? (" + json.dumps(data, ensure_ascii=False) + ")，不正確的話請再呼叫一次。 *[UnieAI專用報價單模板A的檔案下載連結，](" + link_url + ")*"
    response = "請幫我確認你傳入的資料是否正確?? (" + json.dumps(data, ensure_ascii=False) + ")，不正確的話請再呼叫一次。 *[此文件是提供給使用者的檔案，文件下載連結，](" + link_url + ")*"
    wb.save(dst)
    wb.close()
    return response







if __name__ == "__main__":
    main()
