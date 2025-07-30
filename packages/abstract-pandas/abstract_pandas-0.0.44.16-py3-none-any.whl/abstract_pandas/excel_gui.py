from abstract_gui import *
import os
from openpyxl import load_workbook
import pandas as pd
from os.path import isfile,splitext,basename,dirname,isdir,exists,join
MAX_CELLS = 400  # Example limit

from abstract_utilities import make_list
import platform
def save_excel_workbook(file_path):
    try:
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        for wb in excel.Workbooks:
            
            if wb.FullName == file_path:
                wb.Save()
                wb.Close(True)  # True to save changes
        excel.Quit()
    except Exception as e:
        print(f"Error saving Excel file: {e}")
class SingletonMeta(type):
    _instances = {}
    def __call__(cls, *args, **kwargs):
        if cls not in cls._instances:
            cls._instances[cls] = super(SingletonMeta, cls).__call__(*args, **kwargs)
        return cls._instances[cls]
def check_if_excel_file_is_open(file_path):
    try:
        # Attempt to open the file in exclusive mode ('r+' allows reading and writing).
        # The operation will fail if the file is open elsewhere.
        with open(file_path, 'r+', encoding='utf-8'):
            return False
    except Exception as e:
        print(f"The file {file_path} appears to be open in another application. Please close it and try again.")
        return True
def prompt_user_about_open_file(file_path):
    layout = [
        [sg.Text(f'The Excel file {basename(file_path)} appears to be open. Please choose an action:')], 
        [sg.Button('Save and Close'), sg.Button('Close Without Saving'), sg.Button('Cancel')]
    ]
    
    window = sg.Window(f'File Open Warning | {basename(file_path)}', layout)
    
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED or event == 'Cancel':
            break
        elif event == 'Save and Close':
            # Attempt to save and close the file
            save_excel_workbook(file_path)
            print("Saving and closing the file...")
            # Insert logic here to save and close the Excel file
            break
        elif event == 'Close Without Saving':
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Quit()
            # Close the file without saving
            print("Closing the file without saving...")
            # Insert logic here to force close the Excel file
            break
    
    window.close()
def remove_rows_with_queries(df, columns, queries_to_remove):
    """
    Remove rows from a DataFrame where any of a list of queries is found in any of the designated columns.
    
    Parameters:
    - df: pandas.DataFrame to filter.
    - columns: List of column names to search for the queries.
    - queries_to_remove: List of strings or numbers to search for and remove rows if found.
    
    Returns:
    - A modified DataFrame with the specified rows removed.
    """
    if not isinstance(columns, list):
        columns = [columns]
    
    if not isinstance(queries_to_remove, list):
        queries_to_remove = [queries_to_remove]
    
    # Convert all queries to strings
    queries_to_remove = [str(query) for query in queries_to_remove]
    
    # Build a combined mask for all queries
    combined_mask = df[columns].apply(lambda x: x.astype(str)).apply(lambda col: col.apply(lambda cell: not any(query in cell for query in queries_to_remove)))
    
    # Filter the DataFrame based on the combined mask
    filtered_df = df[combined_mask.all(axis=1)]
    
    return filtered_df


def filter_out_rows_with_query(query_string,data,column_names):
    
    mask = [column_names].apply(lambda x: x.str.lower().contains(str(int(query_string)).lower())).any(axis=1)
    
    # Apply the mask to get a DataFrame with only the rows where the query_string is present in at least one of the specified columns
    filtered_df = df[mask]
    return filtered_df


def create_new_name(file_path):
    baseName = basename(file_path)
    dirName = dirname(file_path)
    fileName,ext = splitext(baseName)
    fileNames = [[0] for baseName in os.listdir(dirName) if splitext(baseName)[1] == ext]
    new_name=None
    for i in range(len(fileNames)):
        new_name = f"{fileName}_{i}"
        if new_name not in fileNames:
            file_path = join(dirName,f"{new_name}{ext}")
            break
    return file_path
def read_excel_range(file_path, start_row, num_rows):
    # Skip rows up to start_row, not including the header if it's the first row.
    # Adjust start_row by 1 if your Excel file has headers and you want to include them.
    # start_row is 0-indexed in Python, but 1-indexed in Excel, so adjust accordingly.
    skip = start_row - 1 if start_row > 0 else None
    
    # Read the specified range of rows
    df = pd.read_excel(file_path, skiprows=skip, nrows=num_rows, header=None if skip is None else 0)
    
    return df
def count_rows_in_excel(file_path):
    # Load the workbook in read-only mode
    wb = load_workbook(filename=file_path, read_only=True)
    # Get the first sheet name
    first_sheet_name = wb.sheetnames[0]
    # Select the first sheet
    ws = wb[first_sheet_name]
    # Count the rows in the sheet
    row_count = ws.max_row
    # Close the workbook to free memory
    wb.close()
    return row_count
def read_sample_values(file_path, column_name):
    if file_path and column_name:
        df = pd.read_excel(file_path, usecols=[column_name])
        return df[column_name].dropna().tolist()  # Remove NA values for clean display
    return []
def get_unique_values_from_column(file_path, column_name):
    # Load the specified column of the Excel file into a DataFrame
    df = pd.read_excel(file_path, usecols=[column_name])
    
    # Convert the column to a set of unique values
    unique_values = set(df[column_name].dropna())
    
    return unique_values
def parse_zip_codes(zip_codes):
    sanitized_zips = set()  # Use a set to automatically avoid duplicates
    # Replace various delimiters with space, then split by space
    zip_codes = zip_codes.replace(',', ' ').replace('\t', ' ').replace('\n', ' ').split()
    
    for zip_code in zip_codes:
        # Handle zip code ranges indicated by a dash
        if '-' in zip_code:
            start, end = zip_code.split('-')
            if start.isdigit() and end.isdigit() and len(start) == 5 and len(end) == 5:
                for z in range(int(start), int(end) + 1):
                    sanitized_zips.add(str(z))
        # Handle individual zip codes
        else:
            if zip_code.isdigit() and len(zip_code) == 5:
                sanitized_zips.add(zip_code)
    return list(sanitized_zips)
def open_file_for_client(file_path):
    try:
        if platform.system() == 'Windows':
            os.startfile(file_path)
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', file_path])
        else:  # Assume Linux or other Unix-like
            subprocess.run(['xdg-open', file_path])
    except Exception as e:
        print(f"Failed to open the file {file_path}. Error: {e}")
def keep_rows_with_queries(df, columns, queries_to_keep):
    """
    Keep only the rows in a DataFrame where any of a list of queries is found in any of the designated columns.
    
    Parameters:
    - df: pandas.DataFrame to filter.
    - columns: List of column names to search for the queries.
    - queries_to_keep: List of strings or numbers to search for and keep rows if found.
    
    Returns:
    - A modified DataFrame with only the specified rows kept.
    """
    if not isinstance(columns, list):
        columns = [columns]
    
    if not isinstance(queries_to_keep, list):
        queries_to_keep = [queries_to_keep]
    
    # Convert all queries to strings
    queries_to_keep = [str(query) for query in queries_to_keep]
    
    # Build a combined mask for all queries
    combined_mask = df[columns].apply(lambda x: x.astype(str)).apply(lambda col: col.apply(lambda cell: any(query in cell for query in queries_to_keep)))
    
    # Filter the DataFrame based on the combined mask
    filtered_df = df[combined_mask.any(axis=1)]
    
    return filtered_df
# Function to simulate the displayColumns functionality
def get_excel_columns(file_path):
    df = pd.read_excel(file_path, nrows=0)  # Read only the header to get columns
    return df.columns.tolist()
def get_column_values(file_path,heads=10):
    df= pd.read_excel(file_path, nrows=0)
    return {col: df[col].head(heads).dropna().tolist() for col in df.columns}
def filter_excel(values):
        selected_columns = make_list(values['-COLUMNS-'])
        zip_codes = parse_zip_codes(values["-ZIP_CODES-"])
        action = 'retain' if values['-RETAIN-'] else 'parse out'
        file_path = values['-FILE-']
        new_file_path = values['-NEW_FILE-']
        for path_ in [new_file_path,file_path]:
            if check_if_excel_file_is_open(path_):
                prompt_user_about_open_file(path_)
        df = pd.read_excel(file_path, dtype=str, engine='openpyxl')  # Ensure all data is read as string
        if values['-RETAIN-']:
            df_filtered = keep_rows_with_queries(df,selected_columns,zip_codes)
        if values['-PARSEOUT-']:
            df_filtered = remove_rows_with_queries(df,selected_columns,zip_codes)
        df_filtered.to_excel(new_file_path, index=False, engine='openpyxl')
        sg.popup(f"Submitted!\nSelected Column: {selected_column}\nZip Codes: {zip_codes}\nAction: {action}")
        open_file_for_client(path_mgr.get_path('-NEW_FILE-'))
class pathManager(metaclass=SingletonMeta):
    def __init__(self,window):
        if not hasattr(self, 'initialized'):
            self.initialized = True
            self.path_tracker={}
            self.window=window
    def update_path(self,key,path):
        self.path_tracker[key]=path
    def get_path(self,key):
        return self.path_tracker[key]
    def get_display(self,key):
        return basename(self.get_path(key))
    def create_dir(self,key,file_path):
        if not isdir(dirname(file_path)):
            file_path = join(dirname(self.get_path(key)),file_path)
        self.update_path(key,file_path)
    def display_name(self,values,key):
        if not values['-PATH_CHECK-']:
            self.window[key].update(basename(self.get_path(key)))
        else:
            self.window[key].update(self.get_path(key))
class historyManager(metaclass=SingletonMeta):
    def __init__(self,window):
        if not hasattr(self, 'initialized'):
            self.initialized = True
            self.historyTracker={}
            self.window=window
    def update_history(self,key,value):
        if key not in self.historyTracker:
            self.historyTracker[key]=[]
        self.historyTracker[key].append(value)
    def return_prior_history(self,key):
        if key not in self.historyTracker:
            self.historyTracker[key]=['']
        return self.historyTracker[key][-1]
class excelManager:
    def __init__(self,path_mgr,window,file_path):
        if not hasattr(self, 'initialized'):
            self.initialized = True
            self.path_mgr=path_mgr
            self.dataframeTracker={"filePath":file_path,}
            self.initialize_data(file_path)
            self.window = window
    def get_df(self):
        return self.dataframeTracker["dataFrame"]
    def get_range(self,start_row=0,nrows=100):
        return read_excel_range(self.get_file_path(), start_row=start_row, num_rows=nrows)
    def read_excel(self, start_row=0, nrows=100):
        if start_row:
            df=self.get_range(start_row=start_row, nrows=nrows)
        else:
            df=pd.read_excel(self.get_file_path(),nrows=nrows)
        
        self.dataframeTracker['dataFrame'] = df
        self.dataframeTracker['range'] = {"start_row": start_row, "num_rows": nrows}
        return df
    def get_file_path(self):
        return self.dataframeTracker["filePath"]
    def get_headers(self):
        if "columnHeaders" not in self.dataframeTracker:
            self.dataframeTracker["columnHeaders"] = self.get_df().columns.tolist()
        return self.dataframeTracker["columnHeaders"]
    def get_dict(self):
        return self.dataframeTracker["dataDict"]
    def get_dimensions(self):
        return self.dataframeTracker['dimensions']
    def get_number_rows(self):
        return self.dataframeTracker['rows']
    def get_last_column(self):
        if "last_column" not in self.dataframeTracker:
            self.dataframeTracker['last_column'] = self.dataframeTracker["columnHeaders"][0]
        return self.dataframeTracker["last_column"]
    def read_file(self,file_path=None):
        file_path = file_path or self.path_mgr.get_path('-FILE-') or self.get_file_path()
        if file_path and file_path != self.get_file_path():
            self.initialize_data(file_path)
    def initialize_data(self,file_path):
        self.dataframeTracker["filePath"]=file_path
        self.read_excel()
        self.dataframeTracker['rows']=count_rows_in_excel(self.get_file_path())
        self.dataframeTracker["dataDict"] = self.get_df().to_dict(orient='records')
        self.dataframeTracker["columnHeaders"] = self.get_headers()
        self.dataframeTracker["columnItters"]={}
        for column in self.get_headers():
            self.dataframeTracker["columnItters"][column] = 0
    def update_range_based_on_cell_limit(self, desired_row):
        start_row=self.dataframeTracker['range']['start_row']
        end_row=self.dataframeTracker['range']['num_rows']
        if start_row<=desired_row and desired_row<=end_row:
            return 
        total_columns = len(self.get_headers())
        if total_columns == 0: return  # Prevent division by zero
        
        # Calculate the number of rows we can load based on MAX_CELLS
        num_rows = min(MAX_CELLS // total_columns, self.dataframeTracker["rows"])
        
        # Adjust start_row based on desired_row while keeping within the MAX_CELLS limit
        start_row = max(0, min(desired_row, self.dataframeTracker["rows"] - num_rows))
        
        # Read the new range of rows
        self.read_excel(start_row=start_row, nrows=num_rows)
                    
    def get_values(self,i):
        self.update_range_based_on_cell_limit(i)
        return self.get_df().iloc[i]
    def get_value(self,column,i):
        return self.get_values(i)[column]
    def update_column_itter(self,column,i):
        self.dataframeTracker["columnItters"][column] = i
        return self.get_last_column_itter(column)
    def get_column_itter(self,column):
        return self.dataframeTracker["columnItters"][column]
    def get_last_column_itter(self,column):
        i = self.get_column_itter(column)
        return self.get_value(column,i)
    def get_iter_direction(self,column,k):
        i = self.get_column_itter(column)
        if i+k>0 and i+k<self.get_number_rows():
            i=i+k
            self.update_column_itter(column,i)
        return self.get_value(column,i)
    def update_values_display(self,column=None,k=0):
        column = column or excel_mgr.get_last_column()
        self.dataframeTracker["last_column"] = column
        self.window['-SAMPLE-'].update(f"row:{self.get_column_itter(column)} value: {self.get_iter_direction(column,k)}")
    def get_all_values_from_column(self,column):
        return get_unique_values_from_column(self.get_file_path(),column)
layout = [[sg.Text('Upload Excel File')],
    [sg.Input(key='-FILE-', enable_events=True,disabled=True), sg.FileBrowse('Browse'), sg.Button('Upload')],
    [sg.Input('',key='-NEW_FILE-',enable_events=True),sg.FileBrowse('Browse',enable_events=True,key="-NEW_BROWSE-"),sg.Checkbox('show_path',False,enable_events=True,key="-PATH_CHECK-")],
    [[
    [sg.Frame('select  Zip Code column:',[[sg.Listbox(values=[], enable_events=True, size=(20, 4), key='-COLUMNS-')]]),
     sg.Frame("Addr1",[[sg.Combo([],'',key='-ADDRESS1_SELECT-', size=(10, 1),enable_events=True)]]),
     sg.Frame("Addr2",[[sg.Combo([],'',key='-ADDRESS2_SELECT-', size=(10, 1),enable_events=True)]]),
     sg.Frame("City",[[sg.Combo([],'',key='-CITY_SELECT-', size=(10,1),enable_events=True)]]),
     sg.Frame("State",[[sg.Combo([],'',key='-STATE_SELECT-', size=(10,1),enable_events=True)]]),
     sg.Frame("Zip",[[sg.Combo([],'',key='-ZIP_CODE_SELECT-', size=(10, 1),enable_events=True)]])]],],
          
    [sg.Button('<', key='-PREV-',enable_events=True,), sg.Text('', size=(15, 1), key='-SAMPLE-'), sg.Button('>', key='-NEXT-',enable_events=True,),sg.Button('SELECT', key='-ZIP_SELECT-',enable_events=True)],
    [[sg.Text('Enter Zip Codes separated by commas or new lines...')],sg.MLine(key='-ZIP_CODES-', size=(40, 5), default_text="",enable_events=True)],
    [sg.Radio('Retain Rows', 'RADIO1', default=True, key='-RETAIN-'), sg.Radio('Parse Out Rows', 'RADIO1', key='-PARSEOUT-'),sg.Checkbox('Auto Parse',False,enable_events=True,key="-PARSE_CHECK-")],
    [sg.Button('Submit')]]
def whileit(event,values,window):
        path_mgr = pathManager(window=window)
        history_mgr=historyManager(window=window)
        if values.get('-FILE-'):
            
            excel_mgr = excelManager(window=window,path_mgr=path_mgr,file_path=values['-FILE-'])
        if event == '-FILE-':
            path_mgr.update_path('-FILE-',values['-FILE-'])
            excel_mgr = excelManager(window=window,path_mgr=path_mgr,file_path=values['-FILE-'])
            if values['-NEW_FILE-'] == '':
                file_path = values['-FILE-']
                dirName = dirname(file_path)
                baseName = basename(file_path)
                fileName,ext=splitext(baseName)
                file_path = create_new_name(file_path)
                path_mgr.update_path('-NEW_FILE-',file_path)
                path_mgr.display_name(values,'-NEW_FILE-')
        elif event == '-NEW_FILE-':
            new_name = values['-NEW_FILE-']
            path_mgr.create_dir('-NEW_FILE-',new_name)
            #path_mgr.display_name(values,'-NEW_FILE-')
        elif event == "-PATH_CHECK-":
            path_mgr.display_name(values,'-NEW_FILE-')
        elif event == 'Upload' and values['-FILE-']:
            file_path = values['-FILE-']
            columns = get_excel_columns(file_path)
            window['-COLUMNS-'].update(columns)
            window['-SAMPLE-'].update('')  # Clear sample value display
        elif event == '-COLUMNS-':  # Column selection changed
            if values['-COLUMNS-']:
                selected_column = values['-COLUMNS-'][0]
                excel_mgr.update_values_display(selected_column)
        elif 'SELECT' in event:
            window[event].update(excel_mgr.get_all_values_from_column(excel_mgr.get_last_column()))
        elif event == '-PREV-':
            selected_column = values['-COLUMNS-'][0]
            excel_mgr.update_values_display(excel_mgr.get_last_column(),-1)
        elif event == '-NEXT-':
            excel_mgr.update_values_display(excel_mgr.get_last_column(),1)
        if event == "-ZIP_CODES-":
            prior_zips = history_mgr.return_prior_history("-ZIP_CODES-")
            if values["-PARSE_CHECK-"]:
                history_mgr.update_history("-ZIP_CODES-",values["-ZIP_CODES-"])
                window["-ZIP_CODES-"].update(', '.join(parse_zip_codes(values["-ZIP_CODES-"])))
        elif event == 'Submit':
            headers_js={"address1":values['-ADDR1_SELECT-'],"address2":values['-ADDR2_SELECT-'],"city":values['-CITY_SELECT-'],"state":values['-STATE_SELECT-'],"zip":values['-ZIP_CODE_SELECT-'],"external":""}
            if splitext(values['-NEW_FILE-'])[1]==None:
                new_file=f"{values['-NEW_FILE-']}{splitext(values['-FILE-'])}"
                path_mgr.update_path('-NEW_FILE-',new_file)
            filter_excel(values)


def start_gui():
    window_mgr = AbstractWindowManager()
    window = window_mgr.add_window('Select Zip Code Column', layout,event_handlers=whileit)
    file_path = None
    current_column_values = []
    current_index = 0

    window_mgr.while_window()

start_gui()
