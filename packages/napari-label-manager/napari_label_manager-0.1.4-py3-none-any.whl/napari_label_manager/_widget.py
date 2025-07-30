"""
Napari Label Manager Plugin
A plugin for batch management of label colors and opacity in napari.
"""

import os
import re
import threading

# Add Excel support imports
try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

import napari
import numpy as np
from napari.layers import Labels
from napari.utils import colormaps as cmap
from qtpy.QtCore import Qt, QTimer, Signal
from qtpy.QtGui import QFont
from qtpy.QtWidgets import (
    QCheckBox,
    QComboBox,
    QFileDialog,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSlider,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)


class LabelLayerSelector:
    """A specialized layer selector that only shows Labels layers."""

    def __init__(self, combo_widget, viewer, callback_function=None):
        """
        Initialize the label layer selector.

        Parameters
        ----------
        combo_widget : QComboBox
            The combo box widget to populate with labels layers
        viewer : napari.Viewer
            The napari viewer instance
        callback_function : callable, optional
            Function to call when layer selection changes
        """
        self.combo = combo_widget
        self.viewer = viewer
        self.callback = callback_function

        # Connect to viewer events
        self.viewer.layers.events.inserted.connect(self.update_layers)
        self.viewer.layers.events.removed.connect(self.update_layers)
        self.viewer.layers.events.reordered.connect(self.update_layers)

        # Initial population (before connecting the callback to avoid premature calls)
        self.update_layers()

        # Connect combo box change event after initial population
        if self.callback:
            self.combo.currentTextChanged.connect(self.callback)

    def update_layers(self):
        """Update the combo box with only Labels layers."""
        current_selection = self.combo.currentText()
        self.combo.clear()

        # Get all Labels layers
        label_layers = [
            layer.name
            for layer in self.viewer.layers
            if isinstance(layer, Labels)
        ]

        if label_layers:
            self.combo.addItems(label_layers)
            # Try to maintain previous selection if it still exists
            if current_selection in label_layers:
                self.combo.setCurrentText(current_selection)
        else:
            # Add a placeholder when no labels layers are available
            self.combo.addItem("No Labels layers available")
            self.combo.setEnabled(False)
            return

        self.combo.setEnabled(True)

    def get_current_layer(self):
        """Get the currently selected Labels layer."""
        layer_name = self.combo.currentText()
        if layer_name and layer_name != "No Labels layers available":
            try:
                layer = self.viewer.layers[layer_name]
                if isinstance(layer, Labels):
                    return layer
            except KeyError:
                pass
        return None


class LabelManager(QWidget):
    """Main widget for label management."""

    # Signal emitted when colormap changes
    colormap_changed = Signal(object)

    def __init__(self, napari_viewer: napari.Viewer, parent=None):
        super().__init__(parent)
        self.viewer = napari_viewer
        self.current_layer = None
        self.full_color_dict = {}
        self.background_value = 0
        self.max_labels = 100

        # Performance optimization: cache for layer stats
        self._layer_stats_cache = {}
        self._update_timer = QTimer()
        self._update_timer.setSingleShot(True)
        self._update_timer.timeout.connect(self._delayed_update_layer_info)

        # Initialize layer_selector as None first
        self.layer_selector = None

        self.setup_ui()
        self.connect_signals()

        # Initialize the label layer selector
        self.layer_selector = LabelLayerSelector(
            self.layer_combo, self.viewer, self.on_layer_changed
        )

    def setup_ui(self):
        """Setup the user interface."""
        layout = QVBoxLayout()

        # set width
        self.setMinimumWidth(400)
        # Header
        header = QLabel("Label Manager")
        header.setFont(QFont("Arial", 12, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        # Layer selection
        layer_group = QGroupBox("Layer Selection")
        layer_layout = QVBoxLayout()

        self.layer_combo = QComboBox()
        layer_layout.addWidget(QLabel("Select Label Layer:"))
        layer_layout.addWidget(self.layer_combo)

        # Initialize the specialized layer selector
        self.layer_selector = LabelLayerSelector(
            self.layer_combo, self.viewer, self.on_layer_changed
        )

        layer_group.setLayout(layer_layout)
        layout.addWidget(layer_group)

        # Colormap generation
        colormap_group = QGroupBox("Colormap Generation")
        colormap_layout = QVBoxLayout()

        # Number of colors and seed
        gen_layout = QHBoxLayout()
        gen_layout.addWidget(QLabel("Max Labels:"))
        self.max_labels_spin = QSpinBox()
        self.max_labels_spin.setRange(1, 1000)
        self.max_labels_spin.setValue(self.max_labels)
        gen_layout.addWidget(self.max_labels_spin)

        gen_layout.addWidget(QLabel("Random Seed:"))
        self.seed_spin = QSpinBox()
        self.seed_spin.setRange(0, 100)
        self.seed_spin.setValue(50)
        gen_layout.addWidget(self.seed_spin)

        self.generate_btn = QPushButton("Generate New Colormap")
        self.generate_btn.clicked.connect(self.generate_colormap)
        gen_layout.addWidget(self.generate_btn)

        colormap_layout.addLayout(gen_layout)
        colormap_group.setLayout(colormap_layout)
        layout.addWidget(colormap_group)

        # Batch management
        batch_group = QGroupBox("Batch Label Management")
        batch_layout = QVBoxLayout()

        # Label IDs input
        batch_layout.addWidget(
            QLabel("Label IDs (comma-separated, ranges with '-'):")
        )
        self.label_ids_input = QLineEdit()
        self.label_ids_input.setPlaceholderText("e.g., 1,3,5-10,20,25-30")
        batch_layout.addWidget(self.label_ids_input)

        # Quick presets
        presets_layout = QHBoxLayout()
        presets_layout.addWidget(QLabel("Quick presets:"))
        self.preset_first10_btn = QPushButton("First 10")
        self.preset_first10_btn.clicked.connect(
            lambda: self.set_preset_ids("1-10")
        )
        presets_layout.addWidget(self.preset_first10_btn)

        self.preset_next_btn = QPushButton("Next ID")
        self.preset_next_btn.clicked.connect(self.add_next_id)
        presets_layout.addWidget(self.preset_next_btn)

        self.preset_all_btn = QPushButton("All Current")
        self.preset_all_btn.clicked.connect(self.set_all_current_ids)
        presets_layout.addWidget(self.preset_all_btn)

        batch_layout.addLayout(presets_layout)

        # Opacity controls
        opacity_frame = QFrame()
        opacity_layout = QVBoxLayout()

        # Selected labels opacity
        selected_layout = QHBoxLayout()
        selected_layout.addWidget(QLabel("Selected Labels Opacity:"))
        self.selected_opacity_slider = QSlider(Qt.Horizontal)
        self.selected_opacity_slider.setRange(0, 100)
        self.selected_opacity_slider.setValue(100)
        self.selected_opacity_label = QLabel("1.00")
        self.selected_opacity_slider.valueChanged.connect(
            lambda v: self.selected_opacity_label.setText(f"{v/100:.2f}")
        )
        selected_layout.addWidget(self.selected_opacity_slider)
        selected_layout.addWidget(self.selected_opacity_label)
        opacity_layout.addLayout(selected_layout)

        # Other labels opacity
        other_layout = QHBoxLayout()
        other_layout.addWidget(QLabel("Other Labels Opacity:"))
        self.other_opacity_slider = QSlider(Qt.Horizontal)
        self.other_opacity_slider.setRange(0, 100)
        self.other_opacity_slider.setValue(50)
        self.other_opacity_label = QLabel("0.50")
        self.other_opacity_slider.valueChanged.connect(
            lambda v: self.other_opacity_label.setText(f"{v/100:.2f}")
        )
        other_layout.addWidget(self.other_opacity_slider)
        other_layout.addWidget(self.other_opacity_label)
        opacity_layout.addLayout(other_layout)

        # Hide other labels option
        self.hide_others_checkbox = QCheckBox(
            "Hide Other Labels (opacity = 0)"
        )
        self.hide_others_checkbox.toggled.connect(self.on_hide_others_toggled)
        opacity_layout.addWidget(self.hide_others_checkbox)

        opacity_frame.setLayout(opacity_layout)
        batch_layout.addWidget(opacity_frame)

        # Apply button
        self.apply_btn = QPushButton("Apply Changes")
        self.apply_btn.clicked.connect(self.apply_changes)
        batch_layout.addWidget(self.apply_btn)

        batch_group.setLayout(batch_layout)
        layout.addWidget(batch_group)

        # Label Annotation
        annotation_group = QGroupBox("Label Annotation")
        annotation_layout = QVBoxLayout()

        # Control buttons
        annotation_control_layout = QHBoxLayout()

        # Sheet name input
        annotation_control_layout.addWidget(QLabel("Sheet:"))
        self.sheet_name_input = QLineEdit("Label Annotations")
        self.sheet_name_input.setFixedWidth(120)
        self.sheet_name_input.setToolTip(
            "Excel sheet name for save/load operations"
        )
        annotation_control_layout.addWidget(self.sheet_name_input)

        # Fill range controls
        annotation_control_layout.addWidget(QLabel("Start:"))
        self.annotation_start_input = QLineEdit("1")
        self.annotation_start_input.setFixedWidth(50)
        annotation_control_layout.addWidget(self.annotation_start_input)

        annotation_control_layout.addWidget(QLabel("End:"))
        self.annotation_end_input = QLineEdit("10")
        self.annotation_end_input.setFixedWidth(50)
        annotation_control_layout.addWidget(self.annotation_end_input)

        self.fill_annotation_btn = QPushButton("Fill")
        self.fill_annotation_btn.clicked.connect(self.fill_annotation_range)
        annotation_control_layout.addWidget(self.fill_annotation_btn)

        self.load_current_labels_btn = QPushButton("Current IDs")
        self.load_current_labels_btn.clicked.connect(
            self.load_current_labels_to_annotation
        )
        annotation_control_layout.addWidget(self.load_current_labels_btn)

        # Load Excel button
        self.load_excel_btn = QPushButton("Load")
        self.load_excel_btn.clicked.connect(self.load_excel_to_annotation)
        self.load_excel_btn.setEnabled(EXCEL_AVAILABLE)
        if not EXCEL_AVAILABLE:
            self.load_excel_btn.setToolTip(
                "Install openpyxl to enable Excel import"
            )
        annotation_control_layout.addWidget(self.load_excel_btn)

        annotation_control_layout.addStretch(1)

        # Save button
        self.save_annotation_btn = QPushButton("Save")
        self.save_annotation_btn.clicked.connect(self.save_annotation_to_excel)
        self.save_annotation_btn.setEnabled(EXCEL_AVAILABLE)
        if not EXCEL_AVAILABLE:
            self.save_annotation_btn.setToolTip(
                "Install openpyxl to enable Excel export"
            )
        annotation_control_layout.addWidget(self.save_annotation_btn)

        annotation_layout.addLayout(annotation_control_layout)

        # Annotation table
        self.annotation_table = QTableWidget()
        self.annotation_table.setColumnCount(3)
        self.annotation_table.setHorizontalHeaderLabels(
            ["digital", "biological", "annotation"]
        )

        # Set column resize modes
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.Stretch
        )
        self.annotation_table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.Stretch
        )

        # Allow multiple row selection
        self.annotation_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.annotation_table.setSelectionMode(QTableWidget.ExtendedSelection)

        # Set initial table size and fill with default range
        self.annotation_table.setRowCount(5)
        self.fill_annotation_range()

        annotation_layout.addWidget(self.annotation_table)

        annotation_group.setLayout(annotation_layout)
        layout.addWidget(annotation_group)

        # Status and info
        info_group = QGroupBox("Status & Info")
        info_layout = QVBoxLayout()

        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: green;")
        info_layout.addWidget(self.status_label)

        self.info_text = QTextEdit()
        self.info_text.setMaximumHeight(100)
        self.info_text.setReadOnly(True)
        info_layout.addWidget(self.info_text)

        info_group.setLayout(info_layout)
        layout.addWidget(info_group)

        self.setLayout(layout)

    def connect_signals(self):
        """Connect viewer signals."""
        # Layer selection is now handled by LabelLayerSelector
        self.viewer.layers.events.removed.connect(self.on_layer_removed)

        # Connect to layer events for cache invalidation
        self.viewer.layers.events.changed.connect(
            self.on_layer_properties_changed
        )

    def on_layer_removed(self, event):
        """Handle layer removal to clean up cache."""
        # Clean up cache for removed layers
        removed_layer = event.value
        layer_id = id(removed_layer)
        if layer_id in self._layer_stats_cache:
            del self._layer_stats_cache[layer_id]

    def on_layer_properties_changed(self, event):
        """Handle layer property changes that might affect cache validity."""
        # Clear cache when layer properties change (e.g., time step)
        if hasattr(event, "source") and hasattr(event.source, "current_step"):
            layer_id = id(event.source)
            if layer_id in self._layer_stats_cache:
                # Only clear if this is a time-series change
                del self._layer_stats_cache[layer_id]

    def on_layer_changed(self, layer_name: str):
        """Handle layer selection change."""
        if layer_name and layer_name != "No Labels layers available":
            # Use the layer selector to get the current layer if available
            if self.layer_selector is not None:
                self.current_layer = self.layer_selector.get_current_layer()
            else:
                # Fallback for initialization period
                try:
                    layer = self.viewer.layers[layer_name]
                    if isinstance(layer, Labels):
                        self.current_layer = layer
                    else:
                        self.current_layer = None
                except KeyError:
                    self.current_layer = None

            if self.current_layer is not None:
                self.update_status(f"Selected layer: {layer_name}", "blue")

                # Initialize colormap if needed
                if hasattr(self.current_layer, "colormap"):
                    self.extract_current_colormap()

                # Clear cache for this layer
                layer_id = id(self.current_layer)
                if layer_id in self._layer_stats_cache:
                    del self._layer_stats_cache[layer_id]

                # Delay layer info update to avoid blocking UI
                self._update_timer.start(100)  # 100ms delay
            else:
                self.update_status("Invalid layer selection", "red")
        else:
            self.current_layer = None
            self.update_status("No Labels layer selected", "orange")

    def extract_current_colormap(self):
        """Extract current colormap from the selected layer."""
        if self.current_layer and hasattr(self.current_layer, "colormap"):
            colormap = self.current_layer.colormap
            if hasattr(colormap, "colors"):
                self.full_color_dict = {
                    i + 1: tuple(color)
                    for i, color in enumerate(colormap.colors)
                }
                self.full_color_dict[None] = (0.0, 0.0, 0.0, 0.0)
                if hasattr(colormap, "background_value"):
                    self.background_value = colormap.background_value

    def set_preset_ids(self, preset_type: str):
        """Set preset label IDs."""
        if preset_type == "1-10":
            self.label_ids_input.setText("1-10")

    def add_next_id(self):
        """Add the next ID from the current layer's label list."""
        if not self.current_layer:
            self.update_status("No layer selected", "red")
            return

        # Get current label IDs from the layer
        current_layer_ids = self.get_current_label_ids()
        if not current_layer_ids:
            self.update_status("No labels found in current layer", "orange")
            return

        # Parse currently selected IDs from input
        current_input = self.label_ids_input.text().strip()
        if current_input:
            selected_ids = set(self.parse_label_ids(current_input))
        else:
            selected_ids = set()

        # Find the next ID that's not already selected
        next_id = None

        if selected_ids:
            # select the maximum ID from the currently selected IDs
            max_selected_id = max(selected_ids)

            # Find the next ID in current_layer_ids that's greater than max_selected_id
            for label_id in sorted(current_layer_ids):
                if label_id > max_selected_id and label_id not in selected_ids:
                    next_id = label_id
                    break

            # If that doesn't work, find the smallest missing ID in current_layer_ids
            if next_id is None:
                for label_id in sorted(current_layer_ids):
                    if label_id not in selected_ids:
                        next_id = label_id
                        break
        else:
            # If no IDs are selected, start from the smallest
            for label_id in sorted(current_layer_ids):
                if label_id not in selected_ids:
                    next_id = label_id
                    break

        if next_id is not None:
            # Add to existing selection
            if current_input:
                new_input = f"{current_input},{next_id}"
            else:
                new_input = str(next_id)
            self.label_ids_input.setText(new_input)
            self.update_status(f"Added next ID: {next_id}", "green")
        else:
            self.update_status(
                "All available IDs are already selected", "orange"
            )

    def parse_label_ids(self, ids_string: str) -> list:
        """Parse label IDs from string input using regex."""
        ids = set()
        pattern = (
            r"(\d+)(?:-(\d+))?"  # Matches single IDs or ranges like "1-5"
        )
        matches = re.findall(pattern, ids_string)
        if not ids_string.strip():
            return ids

        for start, end in matches:
            if end:
                # match a range
                ids.update(range(int(start), int(end) + 1))
            else:
                # match a single ID
                ids.add(int(start))

        return sorted(ids)  # Remove duplicates and sort

    def on_hide_others_toggled(self, checked: bool):
        """Handle hide others checkbox toggle."""
        self.other_opacity_slider.setEnabled(not checked)
        if checked:
            self.other_opacity_label.setText("0.00")
        else:
            self.other_opacity_label.setText(
                f"{self.other_opacity_slider.value()/100:.2f}"
            )

    def generate_colormap(self):
        """Generate a new random colormap."""
        self.max_labels = self.max_labels_spin.value()
        seed = self.seed_spin.value() / 100.0

        # Generate colormap
        colormap = self.generate_random_label_colormap(
            self.max_labels,
            background_value=self.background_value,
            random_seed=seed,
        )

        # Convert to color dict
        self.full_color_dict, self.background_value = (
            self.colormap_to_color_dict(colormap)
        )

        self.update_status(
            f"Generated colormap with {self.max_labels} colors", "green"
        )

    def apply_changes(self):
        """Apply opacity changes to selected labels."""
        if not self.current_layer:
            self.update_status("No layer selected", "red")
            return

        # Parse label IDs
        ids_string = self.label_ids_input.text()
        valid_ids = self.parse_label_ids(ids_string)

        if not valid_ids:
            self.update_status("No valid label IDs provided", "red")
            return

        # Get opacity values
        selected_opacity = self.selected_opacity_slider.value() / 100.0
        other_opacity = (
            0.0
            if self.hide_others_checkbox.isChecked()
            else self.other_opacity_slider.value() / 100.0
        )

        # Apply changes
        filtered_color_dict = self.get_filtered_color_dict(
            self.full_color_dict,
            valid_ids,
            selected_opacity=selected_opacity,
            other_opacity=other_opacity,
        )

        # Create and apply new colormap
        new_colormap = self.color_dict_to_color_map(
            filtered_color_dict,
            name=f"batch_managed_{len(valid_ids)}",
            background_value=self.background_value,
        )

        self.current_layer.colormap = new_colormap

        # Update info
        info_text = f"Applied to {len(valid_ids)} labels: {valid_ids[:10]}"
        if len(valid_ids) > 10:
            info_text += f"... (and {len(valid_ids) - 10} more)"
        info_text += f"\nSelected opacity: {selected_opacity:.2f}"
        info_text += f"\nOther opacity: {other_opacity:.2f}"

        self.info_text.setText(info_text)
        self.update_status("Changes applied successfully", "green")

        # Emit signal
        self.colormap_changed.emit(new_colormap)

    def update_status(self, message: str, color: str = "black"):
        """Update status label."""
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {color};")

    # Core colormap functions
    def generate_random_label_colormap(
        self,
        num_colors: int,
        background_value: int = 0,
        random_seed: float = 0.5,
    ):
        """Generate random label colormap."""
        return cmap.label_colormap(num_colors, random_seed, background_value)

    def colormap_to_color_dict(self, colormap):
        """Convert colormap to color dictionary."""
        color_dict = {
            item_id + 1: tuple(color)
            for item_id, color in enumerate(colormap.colors)
        }
        color_dict[None] = (0.0, 0.0, 0.0, 0.0)
        background_value = (
            colormap.background_value
            if hasattr(colormap, "background_value")
            else 0
        )
        return color_dict, background_value

    def get_filtered_color_dict(
        self,
        full_color_dict,
        valid_ids,
        selected_opacity=1.0,
        other_opacity=0.5,
    ):
        """Get filtered color dictionary with batch opacity management."""
        filtered_color_dict = {}

        for key, color in full_color_dict.items():
            if key is None:
                # Keep background unchanged
                filtered_color_dict[key] = color
            elif key in valid_ids:
                # Apply selected opacity to valid IDs
                filtered_color_dict[key] = (
                    color[0],
                    color[1],
                    color[2],
                    selected_opacity,
                )
            else:
                # Apply other opacity to invalid IDs
                filtered_color_dict[key] = (
                    color[0],
                    color[1],
                    color[2],
                    other_opacity,
                )

        return filtered_color_dict

    def color_dict_to_color_map(
        self, color_dict, name="custom", background_value=0
    ):
        """Convert color dictionary to colormap."""
        direct_colormap = cmap.direct_colormap(color_dict)
        direct_colormap.background_value = background_value
        direct_colormap.name = name
        return direct_colormap

    def get_current_label_count(self) -> int:
        """Get the count of unique non-zero labels in the current layer."""
        if not self.current_layer or not hasattr(self.current_layer, "data"):
            return 0

        # Use cache to avoid expensive computation
        layer_id = id(self.current_layer)
        if layer_id in self._layer_stats_cache:
            return self._layer_stats_cache[layer_id]["count"]

        # For large arrays, use sampling for estimation
        data = self._get_current_time_slice(self.current_layer.data)
        # For large arrays, use sparse-aware sampling for estimation
        if data.size > 10_000_000:  # 10M pixels
            try:
                # Step 1: Quick sparse check - find non-zero positions efficiently
                non_zero_mask = data != 0
                non_zero_count = np.count_nonzero(non_zero_mask)

                # If very few non-zero elements, process them all (exact result)
                if (
                    non_zero_count < 100_000
                ):  # Less than 100k non-zero elements
                    non_zero_values = data[non_zero_mask]
                    unique_labels = np.unique(non_zero_values)
                    count = len(unique_labels)

                    # Cache exact result for sparse arrays
                    self._layer_stats_cache[layer_id] = {
                        "count": count,
                        "ids": unique_labels.tolist(),
                        "is_estimate": False,  # This is exact for sparse arrays
                        "sparsity": non_zero_count / data.size,
                        "non_zero_count": non_zero_count,
                        "data_shape": self.current_layer.data.shape,
                    }
                    return count
                # For denser arrays, sample from non-zero positions only
                return self._sample_from_non_zero_positions(
                    data, non_zero_mask, layer_id
                )

            except MemoryError:
                # Fallback to original block sampling if memory is insufficient
                return self._estimate_label_count_sampling(data, layer_id)
        # For smaller arrays, compute exactly
        unique_labels = np.unique(data)
        non_zero_labels = unique_labels[unique_labels != 0]
        count = len(non_zero_labels)

        # Cache the result
        self._layer_stats_cache[layer_id] = {
            "count": count,
            "ids": non_zero_labels.tolist(),
        }
        return count

    def get_current_label_ids(self) -> list:
        """Get list of unique non-zero label IDs in the current layer."""
        if not self.current_layer or not hasattr(self.current_layer, "data"):
            return []

        # Use cache to avoid expensive computation
        layer_id = id(self.current_layer)
        if layer_id in self._layer_stats_cache:
            return self._layer_stats_cache[layer_id]["ids"]

        # For time-series data, only process current time slice
        data = self._get_current_time_slice(self.current_layer.data)

        # For large arrays, use sparse-aware sampling for estimation
        if data.size > 10_000_000:  # 10M pixels
            try:
                # Step 1: Quick sparse check - find non-zero positions efficiently
                non_zero_mask = data != 0
                non_zero_count = np.count_nonzero(non_zero_mask)

                # If very few non-zero elements, process them all (exact result)
                if (
                    non_zero_count < 100_000
                ):  # Less than 100k non-zero elements
                    non_zero_values = data[non_zero_mask]
                    unique_labels = np.unique(non_zero_values)
                    ids = sorted(unique_labels.tolist())

                    # Cache exact result for sparse arrays
                    self._layer_stats_cache[layer_id] = {
                        "count": len(ids),
                        "ids": ids,
                        "is_estimate": False,  # This is exact for sparse arrays
                        "sparsity": non_zero_count / data.size,
                        "non_zero_count": non_zero_count,
                        "data_shape": self.current_layer.data.shape,
                    }
                    return ids

                # For denser arrays, sample from non-zero positions only
                return self._sample_from_non_zero_positions(
                    data, non_zero_mask, layer_id, return_ids=True
                )

            except MemoryError:
                # Fallback to original block sampling if memory is insufficient
                return self._estimate_label_ids_sampling(data, layer_id)

        # For smaller arrays, compute exactly
        unique_labels = np.unique(data)
        non_zero_labels = unique_labels[unique_labels != 0]
        ids = sorted(non_zero_labels.tolist())

        # Cache the result
        self._layer_stats_cache[layer_id] = {"count": len(ids), "ids": ids}
        return ids

    def _sample_from_non_zero_positions(
        self,
        data: np.ndarray,
        non_zero_mask: np.ndarray,
        layer_id: int,
        return_ids: bool = False,
    ):
        """Sample from non-zero positions only for better efficiency."""
        try:
            # Get non-zero positions efficiently
            non_zero_indices = np.where(non_zero_mask)
            total_non_zero = len(non_zero_indices[0])

            # Sample from non-zero positions
            max_sample_size = min(
                50_000, total_non_zero
            )  # Sample at most 50k non-zero positions

            if total_non_zero <= max_sample_size:
                # Use all non-zero values if small enough
                sampled_values = data[non_zero_mask]
            else:
                # Sample indices from non-zero positions
                sample_step = max(1, total_non_zero // max_sample_size)
                if data.ndim == 2:
                    sampled_indices = (
                        non_zero_indices[0][::sample_step][:max_sample_size],
                        non_zero_indices[1][::sample_step][:max_sample_size],
                    )
                    sampled_values = data[sampled_indices]
                elif data.ndim == 3:
                    sampled_indices = (
                        non_zero_indices[0][::sample_step][:max_sample_size],
                        non_zero_indices[1][::sample_step][:max_sample_size],
                        non_zero_indices[2][::sample_step][:max_sample_size],
                    )
                    sampled_values = data[sampled_indices]
                else:
                    # For higher dimensions, use flat indexing
                    flat_indices = np.ravel_multi_index(
                        non_zero_indices, data.shape
                    )
                    sampled_flat_indices = flat_indices[::sample_step][
                        :max_sample_size
                    ]
                    sampled_values = data.flat[sampled_flat_indices]

            # Get unique labels from sampled values
            unique_labels = np.unique(sampled_values)

            sparsity = total_non_zero / data.size

            if return_ids:
                ids = sorted(unique_labels.tolist())
                self._layer_stats_cache[layer_id] = {
                    "count": len(ids),
                    "ids": ids,
                    "is_estimate": total_non_zero > max_sample_size,
                    "sparsity": sparsity,
                    "non_zero_count": total_non_zero,
                    "data_shape": self.current_layer.data.shape,
                }
                return ids
            else:
                count = len(unique_labels)
                self._layer_stats_cache[layer_id] = {
                    "count": count,
                    "ids": unique_labels.tolist(),
                    "is_estimate": total_non_zero > max_sample_size,
                    "sparsity": sparsity,
                    "non_zero_count": total_non_zero,
                    "data_shape": self.current_layer.data.shape,
                }
                return count

        except (MemoryError, ValueError):
            # Fallback to original sampling method
            if return_ids:
                return self._estimate_label_ids_sampling(data, layer_id)
            else:
                return self._estimate_label_count_sampling(data, layer_id)

    def update_layer_info(self):
        """Update layer information display (now optimized for large datasets)."""
        # This method is now handled by _delayed_update_layer_info
        # to prevent blocking the UI thread
        self._delayed_update_layer_info()

    def set_all_current_ids(self):
        """Set all current label IDs in the input field."""
        if label_ids := self.get_current_label_ids():
            ids_string = ",".join(map(str, label_ids))
            self.label_ids_input.setText(ids_string)

            layer_id = id(self.current_layer)
            cache_info = self._layer_stats_cache.get(layer_id, {})
            is_estimate = cache_info.get("is_estimate", False)
            is_minimal = cache_info.get("minimal_sample", False)

            if is_estimate:
                if is_minimal:
                    self.update_status(
                        f"Set ~{len(label_ids)} label IDs (minimal sample - very large dataset)",
                        "orange",
                    )
                else:
                    original_shape = cache_info.get("data_shape", "unknown")
                    self.update_status(
                        f"Set ~{len(label_ids)} estimated label IDs (shape: {original_shape})",
                        "orange",
                    )
            else:
                self.update_status(
                    f"Set {len(label_ids)} current label IDs,\n label IDs list is {label_ids}",
                    "green",
                )
        else:
            self.update_status("No labels found in current layer", "orange")

    def _get_current_time_slice(self, data: np.ndarray) -> np.ndarray:
        """Get the current time slice if this is a time-series dataset."""
        if hasattr(self.current_layer, "current_step") and data.ndim >= 3:
            # This is likely a time-series dataset
            current_step = getattr(
                self.current_layer, "current_step", [0] * (data.ndim - 2)
            )
            if (
                isinstance(current_step, (list, tuple))
                and len(current_step) > 0
            ):
                # Get the first dimension's current step (usually time)
                time_idx = (
                    current_step[0] if current_step[0] < data.shape[0] else 0
                )
                return data[time_idx]
        return data

    def _estimate_label_count_sampling(
        self, data: np.ndarray, layer_id: int
    ) -> int:
        """Estimate label count using memory-efficient sampling for large arrays."""
        # For time-series data, only process current time slice
        data = self._get_current_time_slice(data)

        # Use smaller sample size for extremely large arrays
        max_sample_size = 500_000  # Reduced from 1M
        sample_size = min(
            max_sample_size, max(10_000, data.size // 100)
        )  # At least 10k, at most 1% of data

        try:
            # Use memory-efficient block sampling instead of random indices
            sample = self._block_sample_array(data, sample_size)

            # Get unique labels in sample
            unique_sample = np.unique(sample)
            non_zero_sample = unique_sample[unique_sample != 0]
            estimated_count = len(non_zero_sample)

            # Cache the estimated result (mark as estimate)
            self._layer_stats_cache[layer_id] = {
                "count": estimated_count,
                "ids": non_zero_sample.tolist(),
                "is_estimate": True,
                "data_shape": self.current_layer.data.shape,  # Store original shape
            }
            return estimated_count

        except MemoryError:
            # Fallback to even smaller sample
            return self._minimal_sample_estimation(data, layer_id)

    def _estimate_label_ids_sampling(
        self, data: np.ndarray, layer_id: int
    ) -> list:
        """Estimate label IDs using memory-efficient sampling for large arrays."""
        # For time-series data, only process current time slice
        data = self._get_current_time_slice(data)

        # Use smaller sample size for extremely large arrays
        max_sample_size = 500_000  # Reduced from 1M
        sample_size = min(
            max_sample_size, max(10_000, data.size // 100)
        )  # At least 10k, at most 1% of data

        try:
            # Use memory-efficient block sampling instead of random indices
            sample = self._block_sample_array(data, sample_size)

            # Get unique labels in sample
            unique_sample = np.unique(sample)
            non_zero_sample = unique_sample[unique_sample != 0]
            ids = sorted(non_zero_sample.tolist())

            # Cache the estimated result (mark as estimate)
            self._layer_stats_cache[layer_id] = {
                "count": len(ids),
                "ids": ids,
                "is_estimate": True,
                "data_shape": self.current_layer.data.shape,  # Store original shape
            }
            return ids

        except MemoryError:
            # Fallback to even smaller sample
            return self._minimal_sample_estimation(
                data, layer_id, return_ids=True
            )

    def _block_sample_array(
        self, data: np.ndarray, sample_size: int
    ) -> np.ndarray:
        """Memory-efficient block sampling without creating large index arrays."""
        # Calculate step size for uniform sampling
        total_size = data.size
        step = max(1, total_size // sample_size)

        # Use numpy's advanced indexing with calculated steps
        if data.ndim == 1:
            return data[::step][:sample_size]
        elif data.ndim == 2:
            h, w = data.shape
            h_step = max(1, h // int(np.sqrt(sample_size)))
            w_step = max(1, w // int(np.sqrt(sample_size)))
            return data[::h_step, ::w_step].ravel()[:sample_size]
        else:
            # For higher dimensions, flatten and sample with step
            flat_data = data.ravel()
            return flat_data[::step][:sample_size]

    def _minimal_sample_estimation(
        self, data: np.ndarray, layer_id: int, return_ids: bool = False
    ):
        """Fallback method for extremely large arrays that cause memory errors."""
        try:
            # Use a very small sample size
            sample_size = min(
                50_000, data.size // 1000
            )  # 0.1% of data or 50k max
            sample = self._block_sample_array(data, sample_size)

            unique_sample = np.unique(sample)
            non_zero_sample = unique_sample[unique_sample != 0]

            if return_ids:
                ids = sorted(non_zero_sample.tolist())
                self._layer_stats_cache[layer_id] = {
                    "count": len(ids),
                    "ids": ids,
                    "is_estimate": True,
                    "minimal_sample": True,
                    "data_shape": self.current_layer.data.shape,
                }
                return ids
            else:
                count = len(non_zero_sample)
                self._layer_stats_cache[layer_id] = {
                    "count": count,
                    "ids": non_zero_sample.tolist(),
                    "is_estimate": True,
                    "minimal_sample": True,
                    "data_shape": self.current_layer.data.shape,
                }
                return count

        except (MemoryError, ValueError, RuntimeError) as e:
            # Ultimate fallback - return minimal info
            self._layer_stats_cache[layer_id] = {
                "count": 0,
                "ids": [],
                "is_estimate": True,
                "error": str(e),
                "data_shape": self.current_layer.data.shape,
            }
            return [] if return_ids else 0

    def _delayed_update_layer_info(self):
        """Update layer information in a delayed manner to avoid blocking UI."""
        if not self.current_layer:
            self.info_text.setText("No layer selected")
            return

        # Start background computation
        self._compute_layer_info_async()

    def _compute_layer_info_async(self):
        """Compute layer information asynchronously."""

        def compute_in_background():
            try:
                label_count = self.get_current_label_count()
                layer_id = id(self.current_layer)
                cache_info = self._layer_stats_cache.get(layer_id, {})
                is_estimate = cache_info.get("is_estimate", False)
                is_minimal = cache_info.get("minimal_sample", False)
                data_shape = cache_info.get("data_shape", "unknown")
                sparsity = cache_info.get("sparsity", None)
                non_zero_count = cache_info.get("non_zero_count", None)

                # Prepare info text
                info_text = f"Current layer: {self.current_layer.name}\n"
                info_text += f"Data shape: {data_shape}\n"

                if sparsity is not None:
                    info_text += f"Sparsity: {sparsity:.4f} ({non_zero_count:,} non-zero pixels)\n"

                if is_estimate:
                    if is_minimal:
                        info_text += f"Estimated labels: ~{label_count} (minimal sample - extremely large dataset)\n"
                    else:
                        info_text += f"Estimated labels: ~{label_count} (sampled from non-zero positions)\n"
                else:
                    if (
                        sparsity is not None and sparsity < 0.01
                    ):  # Less than 1% non-zero
                        info_text += f"Total labels: {label_count} (exact - sparse array)\n"
                    else:
                        info_text += f"Total labels: {label_count}\n"

                # Add performance tip for time-series data
                if (
                    isinstance(data_shape, (tuple, list))
                    and len(data_shape) >= 4
                ):
                    info_text += "Tip: Processing current time slice only for performance\n"

                # Add sparsity optimization info
                if (
                    sparsity is not None and sparsity < 0.1
                ):  # Less than 10% non-zero
                    info_text += (
                        "Optimization: Using sparse-aware processing\n"
                    )

                # Update UI in main thread
                QTimer.singleShot(0, lambda: self.info_text.setText(info_text))

            except (
                MemoryError,
                ValueError,
                RuntimeError,
                AttributeError,
            ) as e:
                error_msg = f"Error computing layer info: {str(e)}"
                QTimer.singleShot(
                    0, lambda: self.update_status(error_msg, "red")
                )

        # Run computation in background thread
        thread = threading.Thread(target=compute_in_background, daemon=True)
        thread.start()

    def fill_annotation_range(self):
        """Fill the annotation table with a range of label IDs."""
        try:
            start_num = int(self.annotation_start_input.text())
            end_num = int(self.annotation_end_input.text())
        except ValueError:
            QMessageBox.warning(
                self,
                "Error",
                "Please enter valid integers for start and end numbers.",
            )
            return

        if start_num > end_num:
            QMessageBox.warning(
                self,
                "Error",
                "Start number cannot be greater than end number.",
            )
            return

        num_rows = end_num - start_num + 1

        # Get current selected rows
        selected_ranges = self.annotation_table.selectedRanges()

        if selected_ranges:
            # Fill only selected rows
            for r_range in selected_ranges:
                # Adjust table size if needed
                if r_range.bottomRow() >= self.annotation_table.rowCount():
                    self.annotation_table.setRowCount(r_range.bottomRow() + 1)

                for row_idx in range(
                    r_range.topRow(), r_range.bottomRow() + 1
                ):
                    # Calculate label ID based on position in selection
                    current_num = start_num + (row_idx - r_range.topRow())

                    label_item = QTableWidgetItem(str(current_num))
                    label_item.setTextAlignment(Qt.AlignCenter)
                    self.annotation_table.setItem(row_idx, 0, label_item)

                    # Keep existing data if any
                    if self.annotation_table.item(row_idx, 1) is None:
                        self.annotation_table.setItem(
                            row_idx, 1, QTableWidgetItem("")
                        )
                    if self.annotation_table.item(row_idx, 2) is None:
                        self.annotation_table.setItem(
                            row_idx, 2, QTableWidgetItem("")
                        )
        else:
            # Fill all rows
            current_rows = self.annotation_table.rowCount()
            # Save existing data
            biological_data = []
            annotation_data = []
            for row in range(current_rows):
                biological_item = self.annotation_table.item(row, 1)
                annotation_item = self.annotation_table.item(row, 2)
                biological_data.append(
                    biological_item.text() if biological_item else ""
                )
                annotation_data.append(
                    annotation_item.text() if annotation_item else ""
                )

            # Set new row count
            self.annotation_table.setRowCount(num_rows)

            for row_idx in range(num_rows):
                current_num = start_num + row_idx
                # Set digital (label ID)
                digital_item = QTableWidgetItem(str(current_num))
                digital_item.setTextAlignment(Qt.AlignCenter)
                self.annotation_table.setItem(row_idx, 0, digital_item)

                # Restore biological data if exists
                if row_idx < len(biological_data):
                    biological_item = QTableWidgetItem(
                        biological_data[row_idx]
                    )
                    self.annotation_table.setItem(row_idx, 1, biological_item)
                else:
                    self.annotation_table.setItem(
                        row_idx, 1, QTableWidgetItem("")
                    )

                # Restore annotation data if exists
                if row_idx < len(annotation_data):
                    annotation_item = QTableWidgetItem(
                        annotation_data[row_idx]
                    )
                    self.annotation_table.setItem(row_idx, 2, annotation_item)
                else:
                    self.annotation_table.setItem(
                        row_idx, 2, QTableWidgetItem("")
                    )

    def load_current_labels_to_annotation(self):
        """Load current layer's label IDs into the annotation table."""
        if not self.current_layer:
            QMessageBox.warning(self, "Error", "No label layer selected.")
            return

        try:
            # Get current label IDs
            label_ids = self.get_current_label_ids()

            if not label_ids:
                QMessageBox.information(
                    self, "Info", "No labels found in current layer."
                )
                return

            # Remove background value (0) if present
            if 0 in label_ids:
                label_ids.remove(0)

            # Save existing annotations
            existing_annotations = {}
            existing_biological = {}
            for row in range(self.annotation_table.rowCount()):
                digital_item = self.annotation_table.item(row, 0)
                biological_item = self.annotation_table.item(row, 1)
                annotation_item = self.annotation_table.item(row, 2)
                if digital_item:
                    try:
                        label_id = int(digital_item.text())
                        if biological_item:
                            existing_biological[label_id] = (
                                biological_item.text()
                            )
                        if annotation_item:
                            existing_annotations[label_id] = (
                                annotation_item.text()
                            )
                    except ValueError:
                        continue

            # Set table size and fill with label IDs
            self.annotation_table.setRowCount(len(label_ids))

            for row_idx, label_id in enumerate(sorted(label_ids)):
                # Set digital (label ID)
                digital_item = QTableWidgetItem(str(label_id - 1))
                digital_item.setTextAlignment(Qt.AlignCenter)
                self.annotation_table.setItem(row_idx, 0, digital_item)

                # Set biological (keep existing if available)
                biological_text = existing_biological.get(label_id, "")
                biological_item = QTableWidgetItem(biological_text)
                self.annotation_table.setItem(row_idx, 1, biological_item)

                # Set annotation (keep existing if available)
                annotation_text = existing_annotations.get(label_id, "")
                annotation_item = QTableWidgetItem(annotation_text)
                self.annotation_table.setItem(row_idx, 2, annotation_item)

            self.update_status(
                f"Loaded {len(label_ids)} labels from current layer", "green"
            )

        except (MemoryError, ValueError, RuntimeError, AttributeError) as e:
            QMessageBox.critical(
                self, "Error", f"Failed to load current labels:\n{str(e)}"
            )

    def save_annotation_to_excel(self):
        """Save annotation table to Excel file with custom sheet name and append mode."""
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(
                self,
                "Error",
                "openpyxl library is not installed.\nPlease install it with: pip install openpyxl",
            )
            return

        # Get file path from user
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Annotation to Excel",
            "label_annotations.xlsx",
            "Excel Files (*.xlsx);;All Files (*)",
        )

        if not file_path:
            return

        # Get sheet name from input field
        sheet_name = self.sheet_name_input.text().strip()
        if not sheet_name:
            sheet_name = "Label Annotations"

        try:
            # Check if file exists and load existing workbook or create new one
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                # If sheet already exists, ask user if they want to overwrite
                if sheet_name in wb.sheetnames:
                    reply = QMessageBox.question(
                        self,
                        "Sheet Exists",
                        f"Sheet '{sheet_name}' already exists. Do you want to overwrite it?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No,
                    )
                    if reply == QMessageBox.Yes:
                        # Remove existing sheet
                        wb.remove(wb[sheet_name])
                    else:
                        return

                # Create new sheet
                ws = wb.create_sheet(title=sheet_name)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name

            # Write headers
            ws.append(["digital", "biological", "annotation"])

            # Set header alignment
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=1, column=col_idx).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

            # Write data
            for row in range(self.annotation_table.rowCount()):
                digital_item = self.annotation_table.item(row, 0)
                biological_item = self.annotation_table.item(row, 1)
                annotation_item = self.annotation_table.item(row, 2)

                # Convert digital value from string to number
                digital_value = digital_item.text() if digital_item else ""
                try:
                    # Try to convert to integer first, then float if that fails
                    if digital_value.strip():
                        if "." in digital_value:
                            digital_value = float(digital_value)
                        else:
                            digital_value = int(digital_value)
                    else:
                        digital_value = ""
                except (ValueError, AttributeError):
                    # Keep as string if conversion fails
                    pass

                biological_value = (
                    biological_item.text() if biological_item else ""
                )
                annotation_value = (
                    annotation_item.text() if annotation_item else ""
                )

                ws.append([digital_value, biological_value, annotation_value])

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except (AttributeError, TypeError, ValueError):
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            QMessageBox.information(
                self,
                "Success",
                f"Annotations saved to sheet '{sheet_name}' in:\n{file_path}",
            )
            self.update_status(
                f"Annotations saved to Excel sheet '{sheet_name}'", "green"
            )

        except (OSError, PermissionError, ValueError) as e:
            QMessageBox.critical(
                self, "Error", f"Failed to save Excel file:\n{str(e)}"
            )
            self.update_status("Failed to save annotations", "red")

    def load_excel_to_annotation(self):
        """Load Excel file and populate annotation table using specified sheet name."""
        if not EXCEL_AVAILABLE:
            QMessageBox.critical(
                self,
                "Error",
                "openpyxl and pandas libraries are not installed.\nPlease install them with: pip install openpyxl pandas",
            )
            return

        # Get file path from user
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Excel File",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)",
        )

        if not file_path:
            return

        # Get sheet name from input field
        sheet_name = self.sheet_name_input.text().strip()
        if not sheet_name:
            sheet_name = "Label Annotations"

        try:
            # Check available sheets in the Excel file
            wb = load_workbook(file_path, read_only=True)
            available_sheets = wb.sheetnames
            wb.close()

            # Determine which sheet to read
            sheet_to_read = None
            if sheet_name in available_sheets:
                sheet_to_read = sheet_name
            else:
                # If specified sheet doesn't exist, use the first sheet
                if available_sheets:
                    sheet_to_read = available_sheets[0]
                    QMessageBox.information(
                        self,
                        "Sheet Not Found",
                        f"Sheet '{sheet_name}' not found. Using first sheet '{sheet_to_read}' instead.",
                    )
                else:
                    QMessageBox.warning(
                        self, "Warning", "No sheets found in the Excel file."
                    )
                    return

            # Read Excel file using pandas with specific sheet
            df = pd.read_excel(
                file_path, sheet_name=sheet_to_read, header=None
            )

            if df.empty:
                QMessageBox.warning(
                    self, "Warning", f"The sheet '{sheet_to_read}' is empty."
                )
                return

            # Find the first column with numeric data (ignoring headers)
            numeric_col_idx = None
            start_row = 0

            # Look for the first row that contains numeric data in any column
            for row_idx in range(len(df)):
                for col_idx in range(len(df.columns)):
                    cell_value = df.iloc[row_idx, col_idx]
                    if pd.notna(cell_value):
                        try:
                            # Try to convert to number
                            float(str(cell_value))
                            numeric_col_idx = col_idx
                            start_row = row_idx
                            break
                        except (ValueError, TypeError):
                            continue
                if numeric_col_idx is not None:
                    break

            if numeric_col_idx is None:
                QMessageBox.warning(
                    self,
                    "Warning",
                    f"No numeric data found in sheet '{sheet_to_read}'.",
                )
                return

            # Extract data from the identified starting row
            data_rows = []
            for row_idx in range(start_row, len(df)):
                # Get the numeric value from the identified column
                cell_value = df.iloc[row_idx, numeric_col_idx]
                if pd.notna(cell_value):
                    try:
                        num_value = int(float(str(cell_value)))

                        # Get data from the next two columns (biological and annotation)
                        biological_value = ""
                        annotation_value = ""

                        # Try to get biological value from next column
                        if numeric_col_idx + 1 < len(df.columns):
                            bio_cell = df.iloc[row_idx, numeric_col_idx + 1]
                            if pd.notna(bio_cell):
                                biological_value = str(bio_cell).strip()

                        # Try to get annotation value from column after that
                        if numeric_col_idx + 2 < len(df.columns):
                            ann_cell = df.iloc[row_idx, numeric_col_idx + 2]
                            if pd.notna(ann_cell):
                                annotation_value = str(ann_cell).strip()

                        data_rows.append(
                            (num_value, biological_value, annotation_value)
                        )
                    except (ValueError, TypeError):
                        continue

            if not data_rows:
                QMessageBox.warning(
                    self,
                    "Warning",
                    f"No valid data rows found in sheet '{sheet_to_read}'.",
                )
                return

            # Set table size based on the number of data entries
            self.annotation_table.setRowCount(len(data_rows))

            # Fill the three columns with data from Excel
            for row_idx, (
                digital_value,
                biological_value,
                annotation_value,
            ) in enumerate(data_rows):
                # Column 0: digital (the numeric value from Excel)
                digital_item = QTableWidgetItem(str(digital_value))
                digital_item.setTextAlignment(Qt.AlignCenter)
                self.annotation_table.setItem(row_idx, 0, digital_item)

                # Column 1: biological (from Excel file)
                biological_item = QTableWidgetItem(biological_value)
                self.annotation_table.setItem(row_idx, 1, biological_item)

                # Column 2: annotation (from Excel file)
                annotation_item = QTableWidgetItem(annotation_value)
                self.annotation_table.setItem(row_idx, 2, annotation_item)

            self.update_status(
                f"Loaded {len(data_rows)} entries from sheet '{sheet_to_read}'",
                "green",
            )

            QMessageBox.information(
                self,
                "Success",
                f"Successfully loaded {len(data_rows)} entries from sheet '{sheet_to_read}'.\n"
                f"Found data starting from row {start_row + 1}, column {numeric_col_idx + 1}.\n"
                f"Loaded digital, biological, and annotation data from 3 columns.",
            )

        except ValueError as e:
            QMessageBox.critical(
                self, "Error", f"Failed to load Excel file:\n{str(e)}"
            )
            self.update_status("Failed to load Excel file", "red")
