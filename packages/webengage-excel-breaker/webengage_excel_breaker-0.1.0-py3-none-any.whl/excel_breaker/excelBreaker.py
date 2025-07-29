import os
import sys
import csv
import argparse
from pathlib import Path
from openpyxl import load_workbook

def handle_xlsx(file_path, output_dir):
    try:
        wb = load_workbook(filename=file_path, read_only=True)
    except Exception as e:
        print(f"❌ Failed to load Excel file: {e}")
        return

    for sheet_name in wb.sheetnames:
        print(f"🔄 Processing sheet: {sheet_name}")
        ws = wb[sheet_name]

        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            print(f"⚠️ Skipping empty sheet: {sheet_name}")
            continue

        clean_sheet_name = sheet_name.replace(" ", "_").replace("/", "_")
        output_file = os.path.join(output_dir, f"{clean_sheet_name}.csv")

        with open(output_file, mode="w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)

        print(f"✅ Saved: {output_file}")

def handle_csv(file_path, output_dir):
    print(f"🔄 Processing into CSV file: {file_path.name}")
    output_file = os.path.join(output_dir, f"{file_path.stem}_copy.csv")

    with open(file_path, mode="r", newline="", encoding="utf-8") as infile, \
         open(output_file, mode="w", newline="", encoding="utf-8") as outfile:
        reader = csv.reader(infile)
        writer = csv.writer(outfile)
        for row in reader:
            writer.writerow(row)

    print(f"✅ Outputted CSV to: {output_file}")

def split_excel_or_csv(file_path):
    file_path = Path(file_path)
    if not file_path.exists():
        print(f"❌ File not found: {file_path}")
        sys.exit(1)

    output_dir = file_path.stem
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Output directory: {output_dir}")

    ext = file_path.suffix.lower()

    if ext == ".csv":
        handle_csv(file_path, output_dir)
    elif ext == ".xlsx":
        handle_xlsx(file_path, output_dir)
    else:
        print(f"❌ Unsupported file format: {ext}")
        sys.exit(1)

def main():
    parser = argparse.ArgumentParser(description="Split Excel/CSV file tabs into separate CSVs.")
    subparsers = parser.add_subparsers(dest='command')

    excel_parser = subparsers.add_parser('excel-break', help='Split all Excel/CSV tabs into separate CSV files')
    excel_parser.add_argument('filename', help='Path to Excel or CSV file')

    args = parser.parse_args()

    if args.command == 'excel-break':
        split_excel_or_csv(args.filename)
        print("✅ All Files processed successfully.....!")
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
