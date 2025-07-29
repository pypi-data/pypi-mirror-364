import asyncio
import logging
import os
from pathlib import Path

import aiofiles  # For async file operations
import docx
import fitz  # PyMuPDF
import openpyxl

logger = logging.getLogger(__name__)


class DocumentLoader:
    @staticmethod
    async def load_document(file_path: str) -> str | None:
        _, extension = os.path.splitext(file_path)
        extension = extension.lower()
        content = None
        logger.info("Attempting to load document: %s", file_path)
        try:
            # os.path.exists is sync, consider aiofiles.os.path.exists if available and critical
            if not os.path.exists(file_path):
                logger.error("File not found: %s", file_path)
                raise FileNotFoundError(f"File not found: {file_path}")

            if extension == ".pdf":
                content = await asyncio.to_thread(DocumentLoader._load_pdf_sync, file_path)
            elif extension == ".docx":
                content = await asyncio.to_thread(DocumentLoader._load_docx_sync, file_path)
            elif extension == ".xlsx":
                content = await asyncio.to_thread(DocumentLoader._load_xlsx_sync, file_path)
            elif extension in ['.txt', '.md', '.py', '.js', '.java', '.html', '.css', '.json', '.yaml', '.log']:
                content = await DocumentLoader._load_text_async(file_path)
            else:
                logger.warning(
                    "Unsupported file type: %s for file %s", extension, file_path)
                raise ValueError(f"Unsupported file type: {extension}")

            char_count = len(content) if content else 0
            logger.info(
                "Successfully loaded content from %s (%d characters).", file_path, char_count)
            return content
        except (FileNotFoundError, ValueError) as e:
            logger.error(
                "Error loading document %s: %s", file_path, e, exc_info=True)
            raise  # Re-raise specific exceptions
        except Exception as e:
            logger.error(
                "Error loading document %s: %s", file_path, e, exc_info=True)
            return None

    @staticmethod  # Blocking, to be run in thread
    def _load_pdf_sync(file_path: str) -> str:
        doc = fitz.open(file_path)
        text = "".join([page.get_text() for page in doc])
        doc.close()
        return text

    @staticmethod  # Blocking
    def _load_docx_sync(file_path: str) -> str:
        doc = docx.Document(file_path)
        return "\n".join([paragraph.text for paragraph in doc.paragraphs])

    @staticmethod  # Blocking
    def _load_xlsx_sync(file_path: str) -> str:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"Sheet: {sheet_name}\n")
            for row in sheet.iter_rows():
                row_values = [
                    str(cell.value) if cell.value is not None else "" for cell in row]
                text_parts.append(", ".join(row_values))
            text_parts.append("\n")
        return "\n".join(text_parts)

    @staticmethod
    async def _load_text_async(file_path: str) -> str:
        try:
            async with aiofiles.open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return await f.read()
        except Exception as e:
            logger.error(
                "Error reading text file %s asynchronously: %s", file_path, e, exc_info=True)
            # Fallback to synchronous read if aiofiles fails for some reason (e.g. special file types)
            # This is a defensive measure; ideally aiofiles should handle standard text files.
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f_sync:
                    return f_sync.read()
            except Exception as e_sync:
                logger.error(
                    "Fallback synchronous read for %s also failed: %s", file_path, e_sync, exc_info=True)
                raise  # Re-raise the original or sync error

    @staticmethod
    async def load_from_directory(directory_path: str) -> list[tuple[str, str]]:
        documents_content = []
        path = Path(directory_path)
        if not path.is_dir():
            logger.error("Provided path is not a directory: %s",
                         directory_path, exc_info=True)
            return []

        file_paths_to_load = []
        for file_path_obj in path.rglob('*'):
            if file_path_obj.is_file():
                file_paths_to_load.append(str(file_path_obj))

        # Load documents concurrently
        tasks = [DocumentLoader.load_document(fp) for fp in file_paths_to_load]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        for i, content_or_exc in enumerate(results):
            file_path = file_paths_to_load[i]
            if isinstance(content_or_exc, Exception):
                logger.error(
                    "Failed to load document %s from directory: %s", file_path, content_or_exc)
            elif content_or_exc:
                documents_content.append((file_path, content_or_exc))

        logger.info(
            "Loaded %d documents from directory %s", len(documents_content), directory_path)
        return documents_content
